#!/usr/bin/env python
# coding: utf-8

# In[11]:


import numpy as np
import pandas as pd
from collections import defaultdict
import pulp


# In[12]:


#add the distance collection


# In[13]:


class v8:  
    def __init__(self, municipalities, technologies, method, ind):
        
        self.ind = ind
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            
            #mass of waste from individual home and buildings
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.home = mass['Home']
            self.build = mass['Building']
            self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites, transfer site, and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            
            tech = pd.read_excel(technologies, 'tech', index_col=0)
            #self.GHG = tech['GHG'] #impact of technnology
            self.lb = tech['l.b'] #lower bound constraint
            self.ub = tech['u.b'] #upper bound constraint
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            
            
            #list of possible conventional product 
            c_p_list = pd.read_excel(technologies, 'subs_ratio')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            self.sub_ratio = sub_ratio
            

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            
            ###Transport of coproduct
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            self.I_tr_cop = tr_cop[self.ind]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            self.I_av = av[self.ind]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            self.I_u_cop = u_cop[self.ind]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T]*y[(T,S)]
                if not pd.isna(self.ub[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,T,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,T,S)] = 0
        
        
        #print(trans_I)
        
        #impact of substitution/avoided impact
        self.sub_I = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
                for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
                    self.sub_I[T,cp,conv_p] = self.coprod[T][cp]*self.sub_ratio[cp][conv_p]*self.I_av[conv_p]
        
        self.sub_I_total = {}
        
        for ((T, cp, conv_p), value) in self.sub_I.items():
            if T not in self.sub_I_total:
                self.sub_I_total[T] = 0
            self.sub_I_total[T] += value
            
        for T, total in self.sub_I_total.items():
            print(f"{T} → {total:.2f} {self.unit} avoided")
        
        #impact net of coproduct use:
        self.use_I = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
                for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
                    self.use_I[T,cp,conv_p] = self.coprod[T][cp]*(self.I_u_cop[cp])
        
        self.use_I_total = {}
        
        for ((T, cp, conv_p), val) in self.use_I.items():
            if T not in self.use_I_total:
                self.use_I_total[T] = 0
            self.use_I_total[T] += val
        
        for T, total in self.use_I_total.items():
            print(f"{T} → {total:.2f} {self.unit} related to use of coproducts")

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M, T, S)] + self.I_T[T] - self.sub_I_total[T]+self.use_I_total[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * self.I_tr_cop[cp] * self.dist[mark][S] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                             #   print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        print(f"Impact environnemental total minimal : {pulp.value(model.objective):.2f} {self.unit}")
        
        #for v in model.variables():
         #   print(v.name, "=", v.varValue)


# In[14]:


terr = 'ARC_simple.xlsx'
impact = 'impact.xlsx'
tech = 'tech_ARC_simple.xlsx'


# In[15]:


testv8 = v8(terr, tech, impact, 'Carbon')


# In[16]:


testv8.optimize()


# In[17]:


testv8_water = v8(terr, tech, impact, 'Water')


# In[18]:


testv8_water.optimize()


# In[19]:


#tester si les solutions decentralisées (compostage) produisent des composts qui ne substituent pas des engrais


# In[20]:


terr2 = 'ARC_simple_2.xlsx'
impact2 = 'impact2.xlsx'
tech2 = 'tech_ARC_simple2.xlsx'


# In[21]:


testv8_2 = v8(terr2, tech2, impact2, 'Carbon')


# In[22]:


testv8_2.optimize()


# In[23]:


testv8_2.coprod_list


# In[24]:


#valorisation locale des coproduits issus des technologies decentralisées --> transport coproduit decentralisé = 0 (compost)


# In[25]:


class v9:  
    def __init__(self, municipalities, technologies, method, ind):
        
        self.ind = ind
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            
            #mass of waste from individual home and buildings
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.home = mass['Home']
            self.build = mass['Building']
            self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites, transfer site, and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            
            tech = pd.read_excel(technologies, 'tech', index_col=0)
            #self.GHG = tech['GHG'] #impact of technnology
            self.lb = tech['l.b'] #lower bound constraint
            self.ub = tech['u.b'] #upper bound constraint
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            
            
            #list of possible conventional product 
            c_p_list = pd.read_excel(technologies, 'subs_ratio')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            self.sub_ratio = sub_ratio
            

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            
            ###Transport of coproduct
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            self.I_tr_cop = tr_cop[self.ind]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            self.I_av = av[self.ind]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            self.I_u_cop = u_cop[self.ind]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T]*y[(T,S)]
                if not pd.isna(self.ub[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,T,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,T,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp[(S,cp,mark)] = 0
        
        
        #print(trans_I)
        
        #impact of substitution/avoided impact
        self.sub_I = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
                for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
                    self.sub_I[T,cp,conv_p] = self.coprod[T][cp]*self.sub_ratio[cp][conv_p]*self.I_av[conv_p]
        
        self.sub_I_total = {}
        
        for ((T, cp, conv_p), value) in self.sub_I.items():
            if T not in self.sub_I_total:
                self.sub_I_total[T] = 0
            self.sub_I_total[T] += value
            
        for T, total in self.sub_I_total.items():
            print(f"{T} → {total:.2f} {self.unit} avoided")
        
        #impact net of coproduct use:
        self.use_I = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
                for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
                    self.use_I[T,cp,conv_p] = self.coprod[T][cp]*(self.I_u_cop[cp])
        
        self.use_I_total = {}
        
        for ((T, cp, conv_p), val) in self.use_I.items():
            if T not in self.use_I_total:
                self.use_I_total[T] = 0
            self.use_I_total[T] += val
        
        for T, total in self.use_I_total.items():
            print(f"{T} → {total:.2f} {self.unit} related to use of coproducts")

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
        
        ##distance of coproduct transport
        
                    
            



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M, T, S)] + self.I_T[T] - self.sub_I_total[T]+self.use_I_total[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * trans_I_cp[(S,cp,mark)] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        print(f"Impact environnemental total minimal : {pulp.value(model.objective):.2f} {self.unit}")
        
        #for v in model.variables():
         #   print(v.name, "=", v.varValue)


# In[26]:


terr_sc1 = 'ARC_scenario1.xlsx'


# In[27]:


sc1 = v9(terr_sc1, tech2, impact2, 'Carbon')


# In[28]:


sc1.optimize()


# In[29]:


#Impact technologie = Impact total -- sans transport


# In[30]:


class v10:  
    def __init__(self, municipalities, technologies, method, ind):
        
        self.ind = ind
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            
            #mass of waste from individual home and buildings
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.home = mass['Home']
            self.build = mass['Building']
            self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            
            tech = pd.read_excel(technologies, 'tech', index_col=0)
            #self.GHG = tech['GHG'] #impact of technnology
            self.lb = tech['l.b'] #lower bound constraint
            self.ub = tech['u.b'] #upper bound constraint
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            self.I_tr_cop = tr_cop[self.ind]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            self.I_av = av[self.ind]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            self.I_u_cop = u_cop[self.ind]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T]*y[(T,S)]
                if not pd.isna(self.ub[T]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp[(S,cp,mark)] = 0
        
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
        
        ##distance of coproduct transport
        
                    
            



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M,S)] + self.I_T[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * trans_I_cp[(S,cp,mark)] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")
        
        #print("\n=== Activation des technologies HC ===")
        
        #for S in self.S_list:
         #   print(f"y(HC,{S}) = {y[('HC', S)].varValue}")
               
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        print(f"Impact environnemental total minimal : {pulp.value(model.objective):.2f} {self.unit}")


# In[31]:


impact3 = 'impact3.xlsx'
tech3 = 'tech_ARC_simple3.xlsx'


# In[32]:


sc1_v10 = v10(terr_sc1, tech3, impact3, 'Carbon')


# In[33]:


sc1_v10.optimize()


# In[34]:


#prise en compte de contrainte de capacité : 10% de leur capacité annuelle


# In[35]:


tech3_ub = 'tech_ARC_simple3_ub.xlsx'


# In[36]:


sc1_v10_ub = v10(terr_sc1, tech3_ub, impact3, 'Carbon')


# In[37]:


sc1_v10_ub.optimize()


# In[38]:


sc1_v10_ub_W = v10(terr_sc1, tech3_ub, impact3, 'Water')


# In[39]:


sc1_v10_ub_W.optimize()


# In[40]:


impact3_HC_ss_sub = 'impact3_HC_sans_sub.xlsx'


# In[41]:


#scénario où il n'y a pas de substitution pour le HC


# In[42]:


sc1_v10_ub_HC = v10(terr_sc1, tech3_ub, impact3_HC_ss_sub, 'Carbon')


# In[43]:


sc1_v10_ub_HC.optimize()


# In[44]:


#impacts d'utilisation des coproduits sont tous inclus


# In[45]:


impact4 = 'impact4.xlsx' #impacts d'utilisation des coproduits sont tous inclus


# In[46]:


sc1_cop_use = v10(terr_sc1, tech3_ub, impact4, 'Carbon')


# In[47]:


sc1_cop_use.optimize()


# In[48]:


sc1_cop_use_W = v10(terr_sc1, tech3_ub, impact4, 'Water')


# In[49]:


#contraintes de capacité de technologies existantes différentes de celles à implanter


# In[50]:


class v11:  
    def __init__(self, municipalities, technologies, method, ind):
        
        self.ind = ind
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            
            #mass of waste from individual home and buildings
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.home = mass['Home']
            self.build = mass['Building']
            self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            self.I_tr_cop = tr_cop[self.ind]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            self.I_av = av[self.ind]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            self.I_u_cop = u_cop[self.ind]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp[(S,cp,mark)] = 0
        
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
        
        ##distance of coproduct transport
        
                    
            



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M,S)] + self.I_T[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * trans_I_cp[(S,cp,mark)] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        print(f"Impact environnemental total minimal : {pulp.value(model.objective):.2f} {self.unit}")


# In[51]:


ARC_sc3 = 'ARC_scenario3.xlsx'


# In[52]:


tech_ARC_sc3 = 'tech_ARC_sc3.xlsx'


# In[53]:


sc3 = v11(ARC_sc3, tech_ARC_sc3, impact4, 'Carbon')


# In[54]:


sc3.optimize()


# In[55]:


sc3_W = v11(ARC_sc3, tech_ARC_sc3, impact4, 'Water')


# In[56]:


#ajout du taux de tri


# In[57]:


class v12:  
    def __init__(self, municipalities, technologies, method, ind, sort):
        
        self.ind = ind
        self.sort = sort
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            #mass of waste sorted from individual home and buildings
            #mass = pd.read_excel(municipalities, 'waste_sorted', index_col=0).fillna(0.0)
            #self.home = mass['Home']
            #self.build = mass['Building']
            #self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_inc = self.I_T['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            self.I_tr_cop = tr_cop[self.ind]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            self.I_av = av[self.ind]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            self.I_u_cop = u_cop[self.ind]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp[(S,cp,mark)] = 0
        
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M,S)] + self.I_T[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * trans_I_cp[(S,cp,mark)] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) +  (self.unsort * self.I_inc)
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f} {self.unit}")


# In[58]:


##scenario où la substitution est nulle pour les énergies


# In[59]:


impact5 = 'impact5_energie_decarbonee.xlsx'


# In[60]:


sc3_sust = v12(ARC_sc3, tech_ARC_sc3, impact5, 'Carbon', 0.8)


# In[61]:


sc3_sust.optimize()


# In[62]:


sc3_sust_0_3 = v12(ARC_sc3, tech_ARC_sc3, impact5, 'Carbon', 0.3)


# In[63]:


sc3_sust_0_3.optimize()


# In[64]:


##scénarios où l'on utilise les sites existants qui acceptent les biodéchets + Villers-st-Paul


# In[576]:


ARC_reel = 'ARC_reel.xlsx'
tech_ARC_reel = 'tech_ARC_reel.xlsx'


# In[577]:


#Civic-led pathway


# In[67]:


sc_reel_civic_led = v12(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.8)


# In[68]:


sc_reel_civic_led.optimize()


# In[69]:


sc_reel_civic_led_EQ = v12(ARC_reel, tech_ARC_reel, impact4, 'Total ecosystem quality', 0.8)


# In[70]:


sc_reel_civic_led_EQ.optimize()


# In[71]:


sc_reel_civic_led_HH = v12(ARC_reel, tech_ARC_reel, impact4, 'Total human health', 0.8)


# In[72]:


sc_reel_civic_led_HH.optimize()


# In[73]:


sc_reel_sust = v12(ARC_reel, tech_ARC_reel, impact5, 'Carbon', 0.8)


# In[74]:


sc_reel_sust.optimize()


# In[75]:


##scenario extrême : energie_decarb - population desengagée #mismatch


# In[76]:


impact6 = 'impact6_decarb_deseng.xlsx'


# In[77]:


mismatch = v12(ARC_reel, tech_ARC_reel, impact6, 'Carbon', 0.2)


# In[78]:


mismatch.optimize()


# In[79]:


##sustainable pathway, avec scénario réel


# In[80]:


sust = v12(ARC_reel, tech_ARC_reel, impact5, 'Carbon', 0.8)


# In[81]:


sust.optimize()


# In[82]:


##desengagement total


# In[83]:


impact7 ='impact7_diseng_tot.xlsx'


# In[84]:


diseng_path = v12(ARC_reel, tech_ARC_reel, impact7, 'Carbon', 0.2)


# In[85]:


diseng_path.optimize()


# In[613]:


ARC_reel_ss_inv = 'ARC_reel_sans_inv.xlsx'


# In[612]:


tech_reel_ss_inv = 'tech_ARC_reel_sans_inv.xlsx'


# In[88]:


civic_ss_inv = v12(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Carbon', 0.8)


# In[89]:


civic_ss_inv.optimize()


# In[90]:


civic_ss_inv_EQ = v12(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total ecosystem quality', 0.8)


# In[91]:


civic_ss_inv_EQ.optimize()


# In[92]:


civic_ss_inv_HH = v12(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total human health', 0.8)


# In[93]:


civic_ss_inv_HH.optimize()


# In[94]:


#v13: stockage des résultats - calcul des impacts pour les autres indicateurs


# In[95]:


class v13:  
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            #mass of waste sorted from individual home and buildings
            #mass = pd.read_excel(municipalities, 'waste_sorted', index_col=0).fillna(0.0)
            #self.home = mass['Home']
            #self.build = mass['Building']
            #self.total = mass['Total']
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I[(M,S)] =(self.col_w*self.dist_col[M])+ (self.tr_w * self.dist[M][S]) 
                    else:
                        trans_I[(M,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp[(S,cp,mark)] = 0
        
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model += pulp.lpSum(x[(M, O, T, S)] * (trans_I[(M,S)] + self.I_T[T])                            for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * trans_I_cp[(S,cp,mark)] for T in self.idtech for cp in self.coprod_list for mark in self.market_list 
                     for S in self.S_list)

        model.solve()
        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) +  (self.unsort * self.I_inc)
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f} {self.unit}")
        
    
    def impact_tot(self):
        
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
            else:
                trans1 = 0.0
                trans2 = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T])
            total2 += val * (trans2 + self.I_T2[T])

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        

        
        


# In[96]:


civic_SI_other_ind = v13(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Carbon', 0.8, 'Total ecosystem quality', 'Total human health')


# In[97]:


civic_SI_other_ind.optimize()


# In[98]:


civic_SI_other_ind.impact_tot()


# In[99]:


civic_SI_other_ind_EQ = v13(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total ecosystem quality', 0.8, 'Carbon', 'Total human health')


# In[100]:


civic_SI_other_ind_EQ.optimize()


# In[101]:


civic_SI_other_ind_EQ.impact_tot()


# In[102]:


civic_SI_other_ind_HH = v13(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total human health', 0.8, 'Carbon', 'Total ecosystem quality')


# In[103]:


civic_SI_other_ind_HH.optimize()


# In[104]:


civic_SI_other_ind_HH.impact_tot()


# In[105]:


civic_other_ind_Carb = v13(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.8, 'Total ecosystem quality', 'Total human health')


# In[106]:


civic_other_ind_Carb.optimize()


# In[107]:


civic_other_ind_Carb.impact_tot()


# In[108]:


civic_other_ind_EQ = v13(ARC_reel, tech_ARC_reel, impact4, 'Total ecosystem quality', 0.8, 'Carbon', 'Total human health')


# In[109]:


civic_other_ind_EQ.optimize()


# In[110]:


civic_other_ind_EQ.impact_tot()


# In[111]:


civic_other_ind_HH = v13(ARC_reel, tech_ARC_reel, impact4, 'Total human health', 0.8, 'Carbon', 'Total ecosystem quality')


# In[112]:


civic_other_ind_HH.optimize()


# In[113]:


civic_other_ind_HH.impact_tot()


# In[114]:


#multi-objectif


# In[115]:


class v14:  
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2, W1, W2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.W1 = W1
        self.W2 = W2
        
        self.N1 = 173048 #EQ
        self.N2 = 2.42 #HH
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        
        for T in self.idtech:
            for M in self.mun:
                for S in self.S_list:
                    if self.scale[T] == 'Centralized':
                        trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                        trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S]) 
                    else:
                        trans_I1[(M,S)] = 0
                        trans_I2[(M,S)] = 0
                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for T in self.idtech:
                        if self.scale[T] == 'Centralized':
                            trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                            trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
                        else:
                            trans_I_cp1[(S,cp,mark)] = 0
                            trans_I_cp2[(S,cp,mark)] = 0
        
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*(trans_I1[(M,S)] + self.I_T1[T]))+                                                ((self.W2/self.N2)*(trans_I2[(M,S)] + self.I_T2[T])))                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)]))                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list) 

        model.solve()
        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) +  (self.unsort * ((self.W1/self.N1)*(self.I_inc1)+                                                                      (self.W2/self.N2)*(self.I_inc2)))
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T])
            total2 += val * (trans2 + self.I_T2[T])
            total += val * (trans + self.I_T[T])
    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop2[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        


# In[116]:


test_multi = v14(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.8, 'Total ecosystem quality', 'Total human health',1, 0)


# In[117]:


test_multi.optimize()


# In[118]:


test_multi.impact_tot()


# In[119]:


class v13_debugg:  ##mono-objectif
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2


        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]

            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        #decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
             #               model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I = {}
        
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                #if self.scale[T] == 'Centralized':
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                #else:
                    #trans_I[(M,S)] = 0
                        
                        
        #impact of transport of coproduct
        trans_I_cp = {}
        
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    #for T in self.idtech:
                        #if self.scale[T] == 'Centralized':
                    trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                        #else:
                         #   trans_I_cp[(S,cp,mark)] = 0
                            
        
    
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model += (pulp.lpSum(x[(M, O, T, S)] * ((trans_I[(M, S)] if self.scale[T] == 'Centralized' else 0.0)            + self.I_T[T]) for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)    + pulp.lpSum(z[(T, S, cp, mark)] * (trans_I_cp[(S, cp, mark)] if self.scale[T] == 'Centralized' else 0.0)        for T in self.idtech for S in self.S_list for cp in self.coprod_list for mark in self.market_list)    + (self.unsort * self.I_inc))

        model.writeLP("biowaste_model.lp")

        
        model.solve()
        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective)
        

                                                                    
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T])
            total2 += val * (trans2 + self.I_T2[T])
            total += val * (trans + self.I_T[T])
    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        


# In[120]:


Car_SI = v13_debugg(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health')


# In[121]:


Car_SI.optimize()


# In[122]:


Car_SI.impact_tot()


# In[123]:


EQ_SI = v13_debugg(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total ecosystem quality', 0.6, 'Carbon', 'Total human health')


# In[124]:


EQ_SI.optimize()


# In[125]:


EQ_SI.impact_tot()


# In[126]:


HH_SI = v13_debugg(ARC_reel_ss_inv, tech_reel_ss_inv, impact4, 'Total human health', 0.6, 'Carbon', 'Total ecosystem quality')


# In[127]:


HH_SI.optimize()


# In[128]:


HH_SI.impact_tot()


# In[129]:


Carb_I = v13_debugg(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health')


# In[130]:


Carb_I.optimize()


# In[131]:


Carb_I.impact_tot()


# In[132]:


EQ_I= v13_debugg(ARC_reel, tech_ARC_reel, impact4, 'Total ecosystem quality', 0.6, 'Carbon', 'Total human health')


# In[133]:


EQ_I.optimize()


# In[134]:


EQ_I.impact_tot()


# In[250]:


HH_I= v13_debugg(ARC_reel, tech_ARC_reel, impact4, 'Total human health', 0.6, 'Carbon', 'Total ecosystem quality')


# In[251]:


HH_I.optimize()


# In[252]:


HH_I.impact_tot()


# In[253]:


class v14_debugg: ##bi-objectif (EQ, HH) 
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2, W1, W2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.W1 = W1
        self.W2 = W2
        
        self.N1 = 173048 #EQ
        self.N2 = 2.42 #HH
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            #model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S]) 

                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                    trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T])))                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)         + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)]))                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list)         +  (self.unsort * ((self.W1/self.N1)*(self.I_inc1)+(self.W2/self.N2)*(self.I_inc2)))

        model.solve()
        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        for M in self.mun:
            print(f"\nCommune : {M}")
            for T in self.idtech:
                for O in self.origin:
                    for S in self.S_list:
                        var = x[(M, O, T,S)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T])
            total2 += val * (trans2 + self.I_T2[T])
            total += val * (trans + self.I_T[T])
    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop2[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        


# In[254]:


EQ_I_0_25 = v14_debugg(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.25, 0.75)


# In[255]:


EQ_I_0_25.optimize()


# In[256]:


EQ_I_0_25.impact_tot()


# In[257]:


EQ_I_0_5 = v14_debugg(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.5, 0.5)


# In[258]:


EQ_I_0_5.optimize()


# In[259]:


EQ_I_0_5.impact_tot()


# In[260]:


EQ_I_0_75 = v14_debugg(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.75, 0.25)


# In[261]:


EQ_I_0_75.optimize()


# In[262]:


EQ_I_0_75.impact_tot()


# In[322]:


multi_HH = v14_debugg(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 1)


# In[323]:


multi_HH.optimize()


# In[324]:


multi_HH.impact_tot()


# In[266]:


#TRI-OBJECTIF


# In[595]:


class v15:  
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2, W, W1, W2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.W1 = W1
        self.W2 = W2
        self.W = W
        
        
        #self.N1 = 80809.38 #EQ
        #self.N2 = 1.18 #HH
        #self.N = 39.38 #Carbon
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
        
        self.N = self.tot_mass * self.I_inc
        self.N1 = self.tot_mass * self.I_inc1
        self.N2 = self.tot_mass * self.I_inc2
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
        
        trans_I1_results = pd.DataFrame(list(trans_I1.items()), columns=["trajet", "Values"])
        trans_I1_results.to_excel("trans_EQ.xlsx", index=False)
        
        trans_I2_results = pd.DataFrame(list(trans_I2.items()), columns=["trajet", "Values"])
        trans_I2_results.to_excel("trans_HH.xlsx", index=False)
        
        trans_I_results = pd.DataFrame(list(trans_I.items()), columns=["trajet", "Values"])
        trans_I_results.to_excel("trans_C.xlsx", index=False)
        
        


                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                    trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
                    trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
        
        trans_I_cp1_results = pd.DataFrame(list(trans_I_cp1.items()), columns=["trajet", "Values"])
        trans_I_cp1_results.to_excel("cp_EQ.xlsx", index=False)
        
        trans_I_cp2_results = pd.DataFrame(list(trans_I_cp2.items()), columns=["trajet", "Values"])
        trans_I_cp2_results.to_excel("cp_HH.xlsx", index=False)
        
        trans_I_cp_results = pd.DataFrame(list(trans_I_cp.items()), columns=["trajet", "Values"])
        trans_I_cp_results.to_excel("cp_C.xlsx", index=False)
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
            



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+\
                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)])+
                                           ((self.W/self.N)*trans_I_cp[(S,cp,mark)]))\
                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))))

        model.solve()
        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T])
            total2 += val * (trans2 + self.I_T2[T])
            total += val * (trans + self.I_T[T])
    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table

        
        
        


# In[292]:


I_peat = 'impact_carb_cop.xlsx' #impact évité de la tourbe modifiée pour le compostage individuel et les digestats


# In[293]:


I_tri_1_peat = v15(ARC_reel, tech_ARC_reel, I_peat, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.25, 0.75)


# In[294]:


I_tri_1_peat.optimize()


# In[295]:


I_tri_1_peat.impact_tot()


# In[216]:


test = v15(ARC_reel, tech_ARC_reel, impact4, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.25, 0.75)


# In[217]:


I_tri = 'impact_carb.xlsx'#impact où le C est enlevé du HH et du EQ


# In[218]:


I_tri_1 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.25, 0.75)


# In[219]:


I_tri_1.optimize()


# In[220]:


I_tri_1.impact_tot()


# In[221]:


I_tri_1.flows_all()


# In[222]:


I_tri_1.flows_techno()


# In[223]:


I1 = I_tri_1.flows_df()
I1.to_excel("I1.xlsx",index=False)


# In[663]:


I_tri_1_SI = v15(ARC_reel_ss_inv, tech_reel_ss_inv, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.25, 0.75)


# In[664]:


I_tri_1_SI.optimize()


# In[665]:


I_tri_1_SI.flows_all()


# In[227]:


I_tri_1_SI.flows_techno()


# In[228]:


I_tri_1_SI.impact_tot()


# In[229]:


I_tri_1_SI_peat = v15(ARC_reel_ss_inv, tech_reel_ss_inv, I_peat, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.25, 0.75)


# In[230]:


I_tri_1_SI_peat.optimize()


# In[231]:


I_tri_1_SI_peat.impact_tot()


# In[232]:


I_tri_1_SI_peat.flows_techno()


# In[233]:


I_tri_2 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.75, 0.55)


# In[234]:


I_tri_2.optimize()


# In[235]:


I_tri_2.impact_tot()


# In[236]:


I_tri_3 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0, 0.5, 0.5)


# In[237]:


I_tri_3.optimize()


# In[238]:


I_tri_3.impact_tot()


# In[666]:


I_tri_4 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.25, 0, 0.75)


# In[667]:


I_tri_4.optimize()


# In[668]:


I_tri_4.impact_tot()


# In[669]:


I_tri_4.flows_techno()


# In[670]:


I_tri_4_peat = v15(ARC_reel, tech_ARC_reel, I_peat, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.25, 0, 0.75)


# In[671]:


I_tri_4_peat.optimize()


# In[672]:


I_tri_4_peat.impact_tot()


# In[673]:


I_tri_4_peat.flows_techno()


# In[247]:


I_tri_4_peat = v15(ARC_reel_ss_inv, tech_reel_ss_inv, I_peat, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.25, 0, 0.75)


# In[248]:


I_tri_4_peat.optimize()


# In[249]:


I_tri_4_peat.impact_tot()


# In[250]:


I_tri_4_peat.flows_techno()


# In[674]:


I_tri_5 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.25, 0.75, 0)


# In[675]:


I_tri_5.optimize()


# In[676]:


I_tri_5.impact_tot()


# In[682]:


I_tri_6 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.75, 0, 0.25)


# In[683]:


I_tri_6.optimize()


# In[684]:


I_tri_6.impact_tot()


# In[685]:


I_tri_7 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.75, 0.25, 0)


# In[686]:


I_tri_7.optimize()


# In[687]:


I_tri_7.impact_tot()


# In[688]:


I_tri_8 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.5, 0, 0.5)


# In[689]:


I_tri_8.optimize()


# In[690]:


I_tri_8.impact_tot()


# In[263]:


I_tri_9 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.5, 0.5, 0)


# In[264]:


I_tri_9.optimize()


# In[265]:


I_tri_9.impact_tot()


# In[266]:


I_tri_10 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',1/3, 1/3, 1/3)


# In[267]:


I_tri_10.optimize()


# In[268]:


I_tri_10.impact_tot()


# In[269]:


I_tri_11 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.1, 0.9/2, 0.9/2)


# In[270]:


I_tri_11.optimize()


# In[271]:


I_tri_11.impact_tot()


# In[272]:


I_tri_12 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.05, 0.95/2, 0.95/2)


# In[273]:


I_tri_12.optimize()


# In[274]:


I_tri_12.impact_tot()


# In[275]:


I_tri_13 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.01, 0.99/2, 0.99/2)


# In[276]:


I_tri_13.optimize()


# In[277]:


I_tri_13.impact_tot()


# In[278]:


I_tri_14 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.0, 0.80, 0.20)


# In[279]:


I_tri_14.optimize()


# In[280]:


I_tri_14.impact_tot()


# In[281]:


I_tri_15 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.001, 0.999/2, 0.999/2)


# In[282]:


I_tri_15.optimize()


# In[283]:


I_tri_15.impact_tot()


# In[284]:


I_tri_16 = v15(ARC_reel, tech_ARC_reel, I_tri, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.002, 0.998/2, 0.998/2)


# In[285]:


I_tri_16.optimize()


# In[286]:


I_tri_16.impact_tot()


# In[287]:


I_tri_16.idtech


# In[288]:



import matplotlib.pyplot as plt

# -----------------------------
# PARAMÈTRES À MODIFIER
# -----------------------------
excel_path = "test_diagram.xlsx"
sheet_name = "Feuil1"
scenario_col = "Scenario"

indicator_order = ["Carbon", "EQ", "HH"]  # ordre des axes
dashed_if_contains = "_inv"

# -----------------------------
# LECTURE DES DONNÉES
# -----------------------------
df = pd.read_excel(excel_path, sheet_name=sheet_name)

indicators = indicator_order

for c in indicators:
    df[c] = pd.to_numeric(df[c], errors="coerce")

# -----------------------------
# NORMALISATION (SEULEMENT POUR TRACER LES LIGNES)
# -----------------------------
vals = df[indicators].copy()

vmin = vals.min(axis=0)
vmax = vals.max(axis=0)

# normalisation "meilleur = 1, pire = 0"
vals_norm = (vals - vmin) / (vmax - vmin)
vals_norm = vals_norm.fillna(0.0)

# -----------------------------
# TRACÉ DU GRAPHE
# -----------------------------
x = np.arange(len(indicators))

plt.figure(figsize=(10, 5))

# Tracer les lignes verticales (axes des indicateurs)
for xi in x:
    plt.axvline(x=xi, color="gray", linewidth=0.8)

# Tracer chaque scénario (SUR VALEURS NORMALISÉES)
for i, row in df.iterrows():
    scen = str(row[scenario_col])
    y = vals_norm.iloc[i].values

    linestyle = "--" if dashed_if_contains in scen else "-"
    plt.plot(x, y, marker="o", linestyle=linestyle, label=scen)

    # Afficher les VRAIES VALEURS au-dessus des points
    for xi, ind in zip(x, indicators):
        orig = row[ind]
        plt.text(xi, y[list(x).index(xi)] + 0.02,
                 f"{orig:.2e}",
                 ha="center", va="bottom", fontsize=8)

# -----------------------------
# ÉTIQUETTES DES AXES (VRAIES VALEURS)
# -----------------------------
for xi, ind in zip(x, indicators):
    # étiquette min et max réelles sur chaque axe
    plt.text(xi, -0.05, f"{vmax[ind]:.2e}",
             ha="center", va="top", fontsize=9)
    plt.text(xi, 1.05, f"{vmin[ind]:.2e}",
             ha="center", va="bottom", fontsize=9)

# Mise en forme
plt.xticks(x, indicators)
plt.ylim(-0.1, 1.1)
plt.ylabel("Echelle normalisée (pour le tracé des lignes)")
plt.title("Comparaison des scénarios (coordonnées parallèles)")
plt.legend(loc="center left", bbox_to_anchor=(1.02, 0.5))
plt.tight_layout()
plt.show()


# In[289]:


#contrainte de marché des produits conventionnels


# In[588]:


class v16:  
    def __init__(self, municipalities, technologies, method, ind, sort, ind1, ind2, W, W1, W2):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.W1 = W1
        self.W2 = W2
        self.W = W
        
        
        self.N1 = 80809.38 #EQ
        self.N2 = 1.18 #HH
        self.N = 39.38 #Carbon
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
        
        trans_I1_results = pd.DataFrame(list(trans_I1.items()), columns=["trajet", "Values"])
        trans_I1_results.to_excel("trans_EQ.xlsx", index=False)
        
        trans_I2_results = pd.DataFrame(list(trans_I2.items()), columns=["trajet", "Values"])
        trans_I2_results.to_excel("trans_HH.xlsx", index=False)
        
        trans_I_results = pd.DataFrame(list(trans_I.items()), columns=["trajet", "Values"])
        trans_I_results.to_excel("trans_C.xlsx", index=False)
        
        


                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        trans_I_cp = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                    trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
                    trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
        
        trans_I_cp1_results = pd.DataFrame(list(trans_I_cp1.items()), columns=["trajet", "Values"])
        trans_I_cp1_results.to_excel("cp_EQ.xlsx", index=False)
        
        trans_I_cp2_results = pd.DataFrame(list(trans_I_cp2.items()), columns=["trajet", "Values"])
        trans_I_cp2_results.to_excel("cp_HH.xlsx", index=False)
        
        trans_I_cp_results = pd.DataFrame(list(trans_I_cp.items()), columns=["trajet", "Values"])
        trans_I_cp_results.to_excel("cp_C.xlsx", index=False)
        
        
        #impact of substitution/avoided impact of peat
        self.sub_I = {}
        self.sub_I1 = {}
        self.sub_I2 = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            #for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
           #     for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
            self.sub_I[T] = self.avoid.loc['Peat', T]*(self.I_av['Peat']+ (0.7318*1))
            self.sub_I1[T] = self.avoid.loc['Peat', T]*(self.I_av1['Peat']+ (0.7318*0.1685))
            self.sub_I2[T] = self.avoid.loc['Peat', T]*(self.I_av2['Peat'] +(0.7318*0))

        
        
        
        print("=== Substitution peat coefficients (par tonne de déchets traités) ===")
        for T in self.idtech:
            print(T, "avoid=", self.avoid.at[T,"Peat"] if ("Peat" in self.avoid.columns and T in self.avoid.index) else None,
          "sub_I=", self.sub_I[T], self.sub_I1[T], self.sub_I2[T])
        
        print("avoid index (5):", list(self.avoid.index)[:5])
        print("avoid columns:", list(self.avoid.columns))


        
        
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
        # --- Plafonds de marché (dans les unités des indicateurs) ---
        self.X_CO2  = 0# 10000000 #6.45E+04 # kgCO2e max évités (ton ind "Carbon")
        self.X_PDF  =0# 1e10 # 11369.41689
   # PDF max évités (ton ind1)
        self.X_DALY = 0# 1e10 #3.82E-03
   # DALY max évités (ton ind2)

# évitements bruts
        self.A_CO2_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        self.A_PDF_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I1[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

        self.A_DALY_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I2[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

# variables capées
        self.A_CO2_cap  = pulp.LpVariable("Peat_avoided_CO2_cap",  lowBound=0)
        self.A_PDF_cap  = pulp.LpVariable("Peat_avoided_PDF_cap",  lowBound=0)
        self.A_DALY_cap = pulp.LpVariable("Peat_avoided_DALY_cap", lowBound=0)

# linéarisation du min(., X)
        model += self.A_CO2_cap  <= self.A_CO2_raw
        model += self.A_CO2_cap  <= self.X_CO2

        model += self.A_PDF_cap  <= self.A_PDF_raw
        model += self.A_PDF_cap  <= self.X_PDF

        model += self.A_DALY_cap <= self.A_DALY_raw
        model += self.A_DALY_cap <=self.X_DALY
    



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+\
                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)])+
                                           ((self.W/self.N)*trans_I_cp[(S,cp,mark)]))\
                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))))\
            - ((self.W/self.N)*self.A_CO2_cap) - ((self.W1/self.N1) * self.A_PDF_cap)\
        - ((self.W2/self.N2) * self.A_DALY_cap)

        model.solve()
        
        print("=== Market cap check ===")
        print("CO2  raw:", pulp.value(self.A_CO2_raw), "cap:", pulp.value(self.A_CO2_cap), "X:", self.X_CO2)
        print("PDF  raw:", pulp.value(self.A_PDF_raw), "cap:", pulp.value(self.A_PDF_cap), "X:", self.X_PDF)
        print("DALY raw:", pulp.value(self.A_DALY_raw), "cap:", pulp.value(self.A_DALY_cap), "X:", self.X_DALY)

        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T] )# - self.sub_I1[T])
            total2 += val * (trans2 + self.I_T2[T] ) #- self.sub_I2[T])
            total += val * (trans + self.I_T[T] )# - self.sub_I[T])
        


    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        
            # ---------- Crédit tourbe (capé par le marché) ----------
        A1_raw = sum(val * self.sub_I1[T] for (M, O, T, S), val in self.x_opt.items())
        A2_raw = sum(val * self.sub_I2[T] for (M, O, T, S), val in self.x_opt.items())
        A0_raw = sum(val * self.sub_I[T]  for (M, O, T, S), val in self.x_opt.items())

        A1_cap = min(A1_raw, self.X_PDF)
        A2_cap = min(A2_raw, self.X_DALY)
        A0_cap = min(A0_raw, self.X_CO2)

        total1 -= A1_cap
        total2 -= A2_cap
        total  -= A0_cap
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table

        
        
        


# In[541]:


tech_cop = 'tech_ARC_reel_cop.xlsx'
I_peat_2 = 'impact_carb_peat.xlsx' #IC, ad-b,d,f : impact de la substitution enlevée


# In[542]:


I_tri_13_peat_2 = v16(ARC_reel, tech_cop, I_peat_2, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.01, 0.99/2, 0.99/2)


# In[543]:


I_tri_13_peat_2.optimize()


# In[544]:


I_tri_13_peat_2.avoid


# In[545]:


I_tri_13_peat_2.avoid['IC']['Peat']


# In[546]:


I_tri_13_peat_2.flows_techno()


# In[547]:


I_tri_13_peat_2.impact_tot()


# In[ ]:





# In[189]:


I_tri_13_peat = v16(ARC_reel, tech_cop, I_peat, 'Carbon', 0.6, 'Total ecosystem quality', 'Total human health',0.01, 0.99/2, 0.99/2)


# In[190]:


I_tri_13_peat.optimize()


# In[191]:


I_tri_13_peat.impact_tot()


# In[ ]:


##classe pour centraliser tout : CO2, EQ_sans_CO2, HH_sans_CO2, EQ_avec_CO2, HH_avec_CO2


# In[709]:


class v17:  
    def __init__(self, municipalities, technologies, method, sort, ind, ind1, ind2, ind3, ind4, W, W1, W2, W3, W4,                M_peat):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.ind3 = ind3
        self.ind4 = ind4
        self.W = W
        self.W1 = W1
        self.W2 = W2
        self.W3 = W3
        self.W4 = W4
        self.M_peat = M_peat
        

        
        #current scenario's impact (incineration)

        
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            self.unit3 = imp[self.ind3][0]
            self.unit4 = imp[self.ind4][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            self.tr_w3 = imp[self.ind3][1]
            self.tr_w4 = imp[self.ind4][1]
            
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            self.col_w3 = imp[self.ind3][2]
            self.col_w4 = imp[self.ind4][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            self.I_T3 = tech[self.ind3]
            self.I_T4 = tech[self.ind4]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            self.I_inc3 = self.I_T3['Inc']
            self.I_inc4 = self.I_T4['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            self.I_tr_cop3 = tr_cop[self.ind3]
            self.I_tr_cop4 = tr_cop[self.ind4]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            self.I_av3 = av[self.ind3]
            self.I_av4 = av[self.ind4]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
            self.I_u_cop3 = u_cop[self.ind3]
            self.I_u_cop4 = u_cop[self.ind4]
            
        self.N = self.tot_mass * self.I_inc
        self.N1 = self.tot_mass * self.I_inc1
        self.N2 = self.tot_mass * self.I_inc2
        self.N3 = self.tot_mass * self.I_inc3
        self.N4 = self.tot_mass * self.I_inc4
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        trans_I3 = {}
        trans_I4 = {}
        
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans_I3[(M,S)] = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans_I4[(M,S)] = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
        
        #trans_I1_results = pd.DataFrame(list(trans_I1.items()), columns=["trajet", "Values"])
        #trans_I1_results.to_excel("trans_EQ.xlsx", index=False)
        
        #trans_I2_results = pd.DataFrame(list(trans_I2.items()), columns=["trajet", "Values"])
        #trans_I2_results.to_excel("trans_HH.xlsx", index=False)
        
        #trans_I_results = pd.DataFrame(list(trans_I.items()), columns=["trajet", "Values"])
        #trans_I_results.to_excel("trans_C.xlsx", index=False)
        

                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        trans_I_cp = {}
        trans_I_cp3 = {}
        trans_I_cp4 = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                    trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
                    trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                    trans_I_cp3[(S,cp,mark)] = self.I_tr_cop3[cp] * self.dist[mark][S]
                    trans_I_cp4[(S,cp,mark)] = self.I_tr_cop4[cp] * self.dist[mark][S]
        
        #trans_I_cp1_results = pd.DataFrame(list(trans_I_cp1.items()), columns=["trajet", "Values"])
        #trans_I_cp1_results.to_excel("cp_EQ.xlsx", index=False)
        
        #trans_I_cp2_results = pd.DataFrame(list(trans_I_cp2.items()), columns=["trajet", "Values"])
        #trans_I_cp2_results.to_excel("cp_HH.xlsx", index=False)
        
        #trans_I_cp_results = pd.DataFrame(list(trans_I_cp.items()), columns=["trajet", "Values"])
        #trans_I_cp_results.to_excel("cp_C.xlsx", index=False)
        
        
        #impact of substitution/avoided impact of peat
        self.sub_I = {}
        self.sub_I1 = {}
        self.sub_I2 = {}
        self.sub_I3 = {}
        self.sub_I4 = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            #for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
           #     for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
            self.sub_I[T] = self.avoid.loc['Peat', T]*(self.I_av['Peat']+ (0.7318*1))
            self.sub_I1[T] = self.avoid.loc['Peat', T]*(self.I_av1['Peat']+ (0.7318*0.702543))
            self.sub_I2[T] = self.avoid.loc['Peat', T]*(self.I_av2['Peat'] +(0.7318*7.1373e-6))
            self.sub_I3[T] = self.avoid.loc['Peat', T]*(self.I_av3['Peat'] +(0.7318*0.1685))
            self.sub_I4[T] = self.avoid.loc['Peat', T]*(self.I_av4['Peat'] +(0.7318*0))

        
        
        
        #print("=== Substitution peat coefficients (par tonne de déchets traités) ===")
        #for T in self.idtech:
        #    print(T, "avoid=", self.avoid.at[T,"Peat"] if ("Peat" in self.avoid.columns and T in self.avoid.index) else None,
        #  "sub_I=", self.sub_I[T], self.sub_I1[T], self.sub_I2[T], self.sub_I3[T], self.sub_I4[T])
        
        #print("avoid index (5):", list(self.avoid.index)[:5])
        #print("avoid columns:", list(self.avoid.columns))


        
        
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
        # --- Plafonds de marché (dans les unités des indicateurs) ---
        #self.X_CO2  = 0# 10000000 #6.45E+04 # kgCO2e max évités (ton ind "Carbon")
        #self.X_EQ_ssC  =0# 1e10 # 11369.41689
   # PDF max évités (ton ind1)
        #self.X_HH_ssC = 0# 1e10 #3.82E-03# DALY max évités (ton ind2)
        #self.X_EQ = 0
        #self.X_HH = 0
        
        
        self.X_CO2 = self.M_peat*(self.I_av['Peat']+ (0.7318*1))
        self.X_EQ = self.M_peat*(self.I_av1['Peat']+ (0.7318*0.702543))
        self.X_HH = self.M_peat* (self.I_av2['Peat'] +(0.7318*7.1373e-6))
        self.X_EQssC = self.M_peat* (self.I_av3['Peat'] +(0.7318*0.1685))
        self.X_HHssC = self.M_peat* (self.I_av4['Peat'] +(0.7318*0))
        

# évitements bruts
        
        self.A_CO2_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        self.A_EQ_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I1[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

        self.A_HH_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I2[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        
        self.A_EQssC_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I3[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        
        self.A_HHssC_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I4[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

# variables capées
        self.A_CO2_cap  = pulp.LpVariable("Peat_avoided_CO2_cap",  lowBound=0)
        self.A_EQ_cap = pulp.LpVariable("Peat_avoided_EQ_cap", lowBound=0)
        self.A_HH_cap = pulp.LpVariable("Peat_avoided_HH_cap", lowBound=0)
        self.A_EQssC_cap  = pulp.LpVariable("Peat_avoided_EQssC_cap",  lowBound=0)
        self.A_HHssC_cap = pulp.LpVariable("Peat_avoided_HHssC_cap", lowBound=0)

# linéarisation du min(., X)
        model += self.A_CO2_cap  <= self.A_CO2_raw
        model += self.A_CO2_cap  <= self.X_CO2


        
        model += self.A_EQ_cap <= self.A_EQ_raw
        model += self.A_EQ_cap <=self.X_EQ
        
        model += self.A_HH_cap <= self.A_HH_raw
        model += self.A_HH_cap <=self.X_HH
        
        model += self.A_EQssC_cap  <= self.A_EQssC_raw
        model += self.A_EQssC_cap  <= self.X_EQssC

        model += self.A_HHssC_cap <= self.A_HHssC_raw
        model += self.A_HHssC_cap <=self.X_HHssC
        
    



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T]))+\
                                               ((self.W3/self.N3)*((trans_I3[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T3[T]))+\
                                               ((self.W4/self.N4)*((trans_I4[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T4[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+\
                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)])+\
                                           ((self.W/self.N)*trans_I_cp[(S,cp,mark)])+\
                                           ((self.W3/self.N3)*trans_I_cp3[(S,cp,mark)])+\
                                           ((self.W4/self.N4)*trans_I_cp4[(S,cp,mark)]))\
                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))+\
                          ((self.W3/self.N3)*(self.I_inc3))+((self.W4/self.N4)*(self.I_inc4))))\
            - ((self.W/self.N)*self.A_CO2_cap) - ((self.W1/self.N1) * self.A_EQssC_cap)\
        - ((self.W2/self.N2) * self.A_HHssC_cap) - ((self.W3/self.N3)*self.A_EQ_cap) - ((self.W4/self.N4)*self.A_HH_cap)

        model.solve()
        
        print("=== Market cap check ===")
        print("CO2  raw:", pulp.value(self.A_CO2_raw), "cap:", pulp.value(self.A_CO2_cap), "X:", self.X_CO2)
        print("EQssC  raw:", pulp.value(self.A_EQssC_raw), "cap:", pulp.value(self.A_EQssC_cap), "X:", self.X_EQssC)
        print("HHssC raw:", pulp.value(self.A_HHssC_raw), "cap:", pulp.value(self.A_HHssC_cap), "X:", self.X_HHssC)
        print("EQ raw:", pulp.value(self.A_EQ_raw), "cap:", pulp.value(self.A_EQ_cap), "X:", self.X_EQ)
        print("HH raw:", pulp.value(self.A_HH_raw), "cap:", pulp.value(self.A_HH_cap), "X:", self.X_HH)

        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        total3 = 0
        total4 = 0
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans3 = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans4 = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0
                trans3 = 0
                trans4 = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T] )# - self.sub_I1[T])
            total2 += val * (trans2 + self.I_T2[T] ) #- self.sub_I2[T])
            total += val * (trans + self.I_T[T])
            total3 += val * (trans3 + self.I_T3[T])
            total4 += val * (trans4 + self.I_T4[T])# - self.sub_I[T])
        


    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop[cp] * self.dist[mark][S]
                total3 += val * self.I_tr_cop3[cp] * self.dist[mark][S]
                total4 += val * self.I_tr_cop4[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0
                total3 += 0
                total4 += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        total3 += self.unsort * self.I_inc3
        total4 += self.unsort * self.I_inc4
        
            # ---------- Crédit tourbe (capé par le marché) ----------
        A1_raw = sum(val * self.sub_I1[T] for (M, O, T, S), val in self.x_opt.items())
        A2_raw = sum(val * self.sub_I2[T] for (M, O, T, S), val in self.x_opt.items())
        A0_raw = sum(val * self.sub_I[T]  for (M, O, T, S), val in self.x_opt.items())
        A3_raw = sum(val * self.sub_I3[T]  for (M, O, T, S), val in self.x_opt.items())
        A4_raw = sum(val * self.sub_I4[T]  for (M, O, T, S), val in self.x_opt.items())

        A1_cap = min(A1_raw, self.X_EQssC)
        A2_cap = min(A2_raw, self.X_HHssC)
        A0_cap = min(A0_raw, self.X_CO2)
        A3_cap = min(A3_raw, self.X_EQ)
        A4_cap = min(A4_raw, self.X_HH)
        

        total1 -= A1_cap
        total2 -= A2_cap
        total  -= A0_cap
        total3  -= A3_cap
        total4  -= A4_cap
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        print(f"Impact total {self.ind3} : {total3:.2f} {self.unit3}")
        print(f"Impact total {self.ind4} : {total4:.2f} {self.unit4}")
        
        return {self.ind : total, self.ind1: total1, self.ind2 :total2,                self.ind3 : total3 , self.ind4 : total4}
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table
    
    def get_shares_tech_site(self, decimals=2, tol=1e-9, only_feasible=True):
        """
        Retourne les % de flux par (tech, site) sur le total trié traité.
        """
        ts = {}
        total = 0.0
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            if only_feasible and self.poss_tech[T][S] == 0:
                continue
            total += float(v)
            ts[(T, S)] = ts.get((T, S), 0.0) + float(v)

        if total <= tol:
            return {}

        return {(T, S): round(100.0 * f / total, decimals) for (T, S), f in ts.items()}

        
        
        


# In[604]:


I_brut = 'impact_brut.xlsx'


# In[605]:


sim1 = v17(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 1 , 0, 0, 0, 0, 0)


# In[606]:


sim1.optimize()


# In[607]:


sim1.flows_techno()


# In[608]:


sim1.impact_tot()


# In[694]:


ver = v15(ARC_reel, tech_ARC_reel, I_brut,'Carbon', 0.6, 'EQssC', 'HHssC', 0.5, 0, 0.5)


# In[695]:


ver.optimize()


# In[696]:


ver.impact_tot()


# In[614]:


sim2 = v17(ARC_reel_ss_inv, tech_reel_ss_inv, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 1 , 0, 0, 0, 0, 0)


# In[615]:


sim2.optimize()


# In[616]:


sim2.impact_tot()


# In[617]:


sim2.flows_techno()


# In[621]:


sim3 = v17(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0 , 1, 0, 0, 0, 0)


# In[622]:


sim3.optimize()


# In[628]:


sim3.impact_tot()


# In[629]:


sim3.flows_techno()


# In[630]:


sim4 = v17(ARC_reel_ss_inv, tech_reel_ss_inv, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0 , 1, 0, 0, 0, 0)


# In[631]:


sim4.optimize()


# In[632]:


sim4.impact_tot()


# In[634]:


sim4.flows_techno()


# In[635]:


sim5 = v17(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0 , 0, 1, 0, 0, 0)


# In[636]:


sim5.optimize()


# In[637]:


sim5.impact_tot()


# In[638]:


sim5.flows_techno()


# In[639]:


sim6 = v17(ARC_reel_ss_inv, tech_reel_ss_inv, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0 , 0, 1, 0, 0, 0)


# In[640]:


sim6.optimize()


# In[641]:


sim6.impact_tot()


# In[642]:


sim6.flows_techno()


# In[647]:


sim8 = v17(ARC_reel_ss_inv, tech_reel_ss_inv, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 0)


# In[648]:


sim8.optimize()


# In[649]:


sim8.impact_tot()


# In[650]:


sim8.flows_techno()


# In[652]:


sim7 = v17(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 0)


# In[653]:


sim7.optimize()


# In[654]:


sim7.impact_tot()


# In[655]:


sim7.flows_techno()


# In[656]:


sim9 = v17(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0, 0.5, 0)


# In[657]:


sim9.optimize()


# In[658]:


sim9.impact_tot()


# In[659]:


sim9.flows_techno()


# In[660]:


sim10 = v17(ARC_reel_ss_inv, tech_reel_ss_inv, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0, 0.5, 0)


# In[661]:


sim10.optimize()


# In[662]:


sim10.impact_tot()


# In[697]:




def run_from_excel(municipalities, technologies, method,
                   scenarios_xlsx, sheet="scenarios",
                   ind="Carbon", ind1="EQ", ind2="HH", ind3="EQssC", ind4="HHssC",
                   out_xlsx="results_scenarios.xlsx"):

    sc = pd.read_excel(scenarios_xlsx, sheet_name=sheet)

    # option : ne garder que les lignes à exécuter
    if "run" in sc.columns:
        sc = sc[sc["run"].fillna(1).astype(int) == 1].copy()

    # check colonnes minimales
    required = ["sort", "M_peat", "W", "W1", "W2", "W3", "W4"]
    missing = [c for c in required if c not in sc.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans Excel: {missing}")

    rows = []

    for i, r in sc.iterrows():
        name = r["name"] if "name" in sc.columns else f"scenario_{i}"

        model = v17(
            municipalities=municipalities,
            technologies=technologies,
            method=method,
            sort=float(r["sort"]),
            ind=ind, ind1=ind1, ind2=ind2, ind3=ind3, ind4=ind4,
            W=float(r["W"]), W1=float(r["W1"]), W2=float(r["W2"]), W3=float(r["W3"]), W4=float(r["W4"]),
            M_peat=float(r["M_peat"]),
        )

        model.optimize()

        rows.append({
            "name": name,
            "sort": float(r["sort"]),
            "M_peat": float(r["M_peat"]),
            "W": float(r["W"]), "W1": float(r["W1"]), "W2": float(r["W2"]), "W3": float(r["W3"]), "W4": float(r["W4"]),
            "status": pulp.LpStatus[model.status] if hasattr(model, "status") else "NA",
            "objective": float(model.tot_imp),

            # debug caps (super utile)
            "A_CO2_raw": float(pulp.value(model.A_CO2_raw)),
            "A_CO2_cap": float(pulp.value(model.A_CO2_cap)),
            "X_CO2": float(model.X_CO2),

            "A_EQ_raw": float(pulp.value(model.A_EQ_raw)),
            "A_EQ_cap": float(pulp.value(model.A_EQ_cap)),
            "X_EQ": float(model.X_EQ),

            "A_HH_raw": float(pulp.value(model.A_HH_raw)),
            "A_HH_cap": float(pulp.value(model.A_HH_cap)),
            "X_HH": float(model.X_HH),

            "A_EQssC_raw": float(pulp.value(model.A_EQssC_raw)),
            "A_EQssC_cap": float(pulp.value(model.A_EQssC_cap)),
            "X_EQssC": float(model.X_EQssC),

            "A_HHssC_raw": float(pulp.value(model.A_HHssC_raw)),
            "A_HHssC_cap": float(pulp.value(model.A_HHssC_cap)),
            "X_HHssC": float(model.X_HHssC),
        })

    df = pd.DataFrame(rows)

    # Sauvegarde des résultats
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        sc.to_excel(writer, sheet_name="inputs_used", index=False)
        df.to_excel(writer, sheet_name="results", index=False)

    return df


# In[712]:





# In[698]:


df = run_from_excel(
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_brut,
    scenarios_xlsx="test_run.xlsx",
    sheet="Feuil1",
    ind="Carbon",
    ind1="EQ",
    ind2="HH",
    ind3="EQssC",
    ind4="HHssC",
    out_xlsx="results_scenarios.xlsx"
)
print(df.head())


# In[966]:




def run_batch_from_excel(scenarios_xlsx,
    municipalities,
    technologies,
    method, sort,
    out_xlsx="results_batch.xlsx",
    sheet="scenarios"):
    
    scen = pd.read_excel(scenarios_xlsx, sheet_name=sheet)
    
    sort = sort

    rows = []
    for _, r in scen.iterrows():
        name = str(r["name"])
        #sort = float(r["sort"])
        M_peat = float(r["M_peat"])

        # poids (W) : adapte si tes colonnes sont nommées autrement
        W  = float(r["W"])
        W1 = float(r["W1"])
        W2 = float(r["W2"])
        W3 = float(r["W3"])
        W4 = float(r["W4"])

        m = v17(
            municipalities=municipalities,
            technologies=technologies,
            method=method,
            sort=sort,
            ind="Carbon",
            ind1="EQ",
            ind2="HH",
            ind3="EQssC",
            ind4="HHssC",
            W=W, W1=W1, W2=W2, W3=W3, W4=W4,
            M_peat=M_peat
        )
        m.optimize()

        impacts = m.impact_tot()
        shares = m.get_shares_tech_site()

        rows.append({
            "name": name,
            "sort": sort,
            "impacts": impacts,
            "shares": shares,
            "idtech": list(m.idtech),
            "S_list": list(m.S_list),
            "poss_tech": m.poss_tech,
        })

    # construire tableau résultat au format techno/site (2 lignes d'en-têtes)
    df = build_output_table(rows)
    export_two_header_excel(df, out_xlsx)
    return df


# In[967]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

IMPACT_ORDER = ["Carbon", "EQ", "HH", "EQssC", "HHssC"]

def build_output_table(rows):
    # on prend la structure techno/site depuis le 1er modèle (identique pour tous scénarios normalement)
    idtech = rows[0]["idtech"]
    S_list = rows[0]["S_list"]
    poss_tech = rows[0]["poss_tech"]

    # colonnes techno/site faisables
    tech_site_cols = []
    for T in idtech:
        for S in S_list:
            try:
                feasible = float(poss_tech[T][S]) != 0
            except Exception:
                feasible = True
            if feasible:
                tech_site_cols.append((T, S))

    columns = [("", "name"), ("", "sort")] + tech_site_cols + [("", c) for c in IMPACT_ORDER]
    cols = pd.MultiIndex.from_tuples(columns)

    data = []
    for r in rows:
        row = {("", "name"): r["name"], ("", "sort"): r["sort"]}

        # % flux
        for (T, S) in tech_site_cols:
            row[(T, S)] = r["shares"].get((T, S), None)

        # impacts
        for c in IMPACT_ORDER:
            row[("", c)] = r["impacts"].get(c, None)

        data.append(row)

    return pd.DataFrame(data, columns=cols)

def export_two_header_excel(df, out_xlsx):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="results", index=True)

    wb = load_workbook(out_xlsx)
    ws = wb["results"]

    # aligner en-têtes
    for cell in ws[1] + ws[2]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # fusionner techno (ligne 1) sur plusieurs sites
    max_col = ws.max_column
    cur = ws.cell(row=1, column=1).value
    start = 1
    for col in range(2, max_col + 2):
        val = ws.cell(row=1, column=col).value if col <= max_col else None
        if val != cur:
            if cur not in (None, "") and (col - start) > 1:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col-1)
            start = col
            cur = val

    # format % sur colonnes techno/site (top != "" et sub != "")
    for col in range(1, max_col + 1):
        top = ws.cell(row=1, column=col).value
        sub = ws.cell(row=2, column=col).value
        if top not in (None, "") and sub not in (None, ""):
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    # format scientifique sur impacts (top=="" et sub in IMPACT_ORDER)
    for col in range(1, max_col + 1):
        top = ws.cell(row=1, column=col).value
        sub = ws.cell(row=2, column=col).value
        if (top in (None, "")) and (sub in IMPACT_ORDER):
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00E+00"

    wb.save(out_xlsx)


# In[874]:


test1 = run_batch_from_excel(scenarios_xlsx="test_run.xlsx",
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_brut,
    out_xlsx="results_scenarios1.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[939]:


brut_inv = run_batch_from_excel(scenarios_xlsx="brut_invest2.xlsx",
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_brut,
    out_xlsx="brut_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[969]:


brut_inv_sort = run_batch_from_excel(scenarios_xlsx="brut_invest2.xlsx",
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_brut, sort=0.6,
    out_xlsx="brut_inv_result_vf_sort.xlsx", sheet = 'Feuil1')


# In[940]:


brut_ss_inv = run_batch_from_excel(scenarios_xlsx="brut_invest2.xlsx",
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_brut,
    out_xlsx="brut_ss_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[941]:


I_hc_correct = 'impact_HC_correct.xlsx'


# In[942]:


HC_correct_inv = run_batch_from_excel(scenarios_xlsx="brut_invest2.xlsx",
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_hc_correct,
    out_xlsx="HC_corr_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[943]:


HC_correct_ss_inv = run_batch_from_excel(scenarios_xlsx="brut_invest2.xlsx",
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_hc_correct,
    out_xlsx="HC_corr_ss_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[878]:


peat_corr = 'sim_peat_corr.xlsx'
I_peat_corr = 'impact_peat_corr.xlsx'


# In[948]:


peat_corr_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_corr,
    out_xlsx="peat_corr_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[949]:


peat_corr_ss_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_corr,
    out_xlsx="peat_corr_ss_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[881]:


I_peat_HC_corr = 'impact_peat_HC_corr.xlsx'


# In[947]:


peat_corr2 = 'sim_peat_corr2.xlsx'


# In[950]:


peat_HC_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr,
    out_xlsx="peat_HC_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[951]:


peat_HC_ss_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr,
    out_xlsx="peat_HC_ss_inv_result_vf.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[970]:


##analyse de sensibilité par rapport au taux de tri


# In[ ]:


#AS 90%taux de tri


# In[976]:


ASpos_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr, sort = 0.9,
    out_xlsx="ASpos_inv.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[973]:


ASpos = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr, sort = 0.9,
    out_xlsx="ASpos.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[ ]:


##AS taux de tri 30%


# In[975]:


ASneg_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr, sort = 0.3,
    out_xlsx="ASneg_inv.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[977]:


ASneg = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr, sort = 0.3,
    out_xlsx="ASneg.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[1012]:


result_data ='result_norm.xlsx'
result_data2 ='result_norm_vf.xlsx'


# In[1001]:


import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import re
from matplotlib.lines import Line2D
import matplotlib.patheffects as pe
import textwrap

def resultV9(data, sheet, title_of_graph,
             overlap_eps=1e-6,   # tolérance: 0.0 strict, 1e-6/1e-4 si arrondis
             wrap_chars=36,      # largeur max avant retour ligne dans l'encadré
             min_gap_axes=0.08   # espacement vertical entre encadrés (coords axes)
            ):

    scenario_col = "Scenario"
    indicators = ["Carbon", "EQ", "HH", "EQ (excl. CO2)", "HH (excl. CO2)"]
    inv_tag = "_inv"

    df = pd.read_excel(data, sheet)

    # conversion numérique
    for c in indicators:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # --- tri naturel (même si pas de chiffres dans les noms)
    def extract_base(name):
        return str(name).replace(inv_tag, "")

    def scen_sort_key(name):
        name = str(name)
        base = extract_base(name)
        m = re.search(r"\d+", base)
        num = int(m.group()) if m else 10**9
        is_inv = inv_tag in name
        return (num, is_inv, base, name)

    df = df.sort_values(
        by=scenario_col,
        key=lambda col: col.map(scen_sort_key)
    ).reset_index(drop=True)

    # --- normalisation APRES tri
    vals = df[indicators].copy()
    vmin = vals.min(axis=0)
    vmax = vals.max(axis=0)
    denom = (vmax - vmin).replace(0, np.nan)
    vals_norm = ((vals - vmin) / denom).fillna(0.0)

    # --- liste des scénarios "base" pour couleurs (mais ici tes scénarios ne sont pas numérotés,
    # donc base == scénarios eux-mêmes si pas de _inv)
    base_scenarios =  [
    "minCO2",
    "minEQ",
    "minHH",
    "minEQ&HH",
    "minCO2&EQ",
    "minCO2&HH",
    "minCO2&EQ&HH"
]


    # --- palette forte
    distinct_13 = [
        "#1f77b4",  # bleu
        "#d62728",  # rouge
        "#2ca02c",  # vert
        "#ff7f0e",  # orange
        "#9467bd",  # violet
        "#8c564b",  # brun
        "#17becf",  # cyan
        "#bcbd22",  # olive
        "#e377c2",  # rose
        "#7f7f7f",  # gris
        "#444444",  # anthracite (plus doux que noir)
        "#a65628",  # marron/orangé
        "#c2b280",  # beige foncé
    ]
    color_map = {base: distinct_13[i % len(distinct_13)] for i, base in enumerate(base_scenarios)}

    # --- interpolation
    k = 4
    x_base = np.arange(len(indicators))
    x_dense = np.linspace(x_base.min(), x_base.max(), (len(indicators) - 1) * k + 1)
    mark_positions = np.arange(0, len(x_dense), k)  # marqueurs uniquement sur les axes

    fig, ax = plt.subplots(figsize=(11, 5))

    for xi in x_base:
        ax.axvline(x=xi, color="gray", linewidth=0.8, alpha=0.35, zorder=0)

    # paramètres tracé
    lw = 1.25
    a = 1.0
    ms = 6.2
    mew = 1.6

    # ==============================
    # 1) Tracé de toutes les courbes
    # ==============================
    for i, row in df.iterrows():
        scen = str(row[scenario_col])
        base = extract_base(scen)

        y_base = vals_norm.iloc[i].values
        y_dense = np.interp(x_dense, x_base, y_base)

        is_inv = inv_tag in scen
        marker_style = "x" if is_inv else "o"
        line_style = "--" if is_inv else "-"

        line, = ax.plot(
            x_dense, y_dense,
            linestyle=line_style,
            color=color_map[base],
            linewidth=lw,
            alpha=a,
            marker=marker_style,
            markersize=ms,
            markeredgewidth=mew,
            markeredgecolor="black",
            markerfacecolor=color_map[base] if marker_style == "o" else None,
            markevery=mark_positions,
            zorder=2
        )

        line.set_path_effects([
            pe.Stroke(linewidth=lw + 2.2, foreground="white", alpha=0.9),
            pe.Normal()
        ])

    # ==============================================
    # 2) Détection des courbes superposées
    #    On compare les vecteurs normalisés (5 valeurs)
    # ==============================================
    Y = vals_norm.values
    n = Y.shape[0]

    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            if np.max(np.abs(Y[i] - Y[j])) <= overlap_eps:
                adj[i].append(j)
                adj[j].append(i)

    # composantes connexes = groupes superposés
    seen = [False] * n
    groups = []
    for i in range(n):
        if seen[i]:
            continue
        stack = [i]
        comp = []
        seen[i] = True
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in adj[u]:
                if not seen[v]:
                    seen[v] = True
                    stack.append(v)
        if len(comp) >= 2:
            groups.append(comp)

    # =========================================================
    # 3) Annotations avec flèche pour les groupes superposés
    #    -> on pointe vers la zone de gauche (x=0)
    #    -> encadrés empilés pour ne pas se chevaucher
    # =========================================================
    # =========================================================
# 3) Annotations alignées (flèches même longueur)
# =========================================================

    if groups:

        fig.subplots_adjust(left=0.30)

    # Position horizontale fixe des encadrés (même colonne)
        x_box_ax = -0.08   # colonne fixe (coord axes)
        x_target = x_base[0]  # on pointe vers le premier axe vertical

    # Trier du haut vers le bas
        def group_y_left(g):
            return float(np.mean([Y[idx][0] for idx in g]))

        groups_sorted = sorted(groups, key=group_y_left, reverse=True)

    # Espacement vertical régulier
        n_groups = len(groups_sorted)
        y_positions = np.linspace(0.85, 0.15, n_groups)

        for g, y_box in zip(groups_sorted, y_positions):

            scen_list = [str(df.loc[idx, scenario_col]) for idx in g]
            scen_list = sorted(scen_list, key=scen_sort_key)

            text = ", ".join(scen_list)
            text = "\n".join(textwrap.wrap(text, width=36, break_long_words=False))

            ax.annotate(
            text,
            xy=(x_target, group_y_left(g)),
            xycoords="data",
            xytext=(x_box_ax, y_box),
            textcoords=ax.transAxes,
            ha="right",   # ← alignement côté droit
            va="center",
            fontsize=8,
            color="#333333",
            bbox=dict(
                boxstyle="round,pad=0.28",
                fc="white",
                ec="#222222",
                lw=1.1,
                alpha=0.98
            ),
            arrowprops=dict(
                arrowstyle="-|>",
                lw=1.8,
                color="#333333",
                mutation_scale=18,
                shrinkA=5,
                shrinkB=5
            ),
            clip_on=False,
            zorder=10)


    # --- valeurs affichées (haut=max, bas=min) comme ton code
    for xi, ind in zip(x_base, indicators):
        ax.text(xi - 0.07, 1.02, f"{vmax[ind]:.2e}",
                ha="right", va="bottom", fontsize=8, alpha=0.85)
        ax.text(xi - 0.07, -0.02, f"{vmin[ind]:.2e}",
                ha="right", va="top", fontsize=8, alpha=0.85)

    ax.set_xticks(x_base)
    ax.set_xticklabels(indicators)
    ax.set_ylim(-0.1, 1.1)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.grid(False)
    ax.set_title(title_of_graph)

    # --- légendes (comme avant)
    color_handles = [
        Line2D([0], [0], color=color_map[base], lw=2, label=base)
        for base in base_scenarios
    ]

    style_handles = [
        Line2D([0], [0], color="black", lw=lw, marker="o",
               markersize=6, markeredgewidth=1.4, markeredgecolor="black",
               linestyle="-", label="No investment"),
        Line2D([0], [0], color="black", lw=lw, marker="x",
               markersize=6, markeredgewidth=1.6, markeredgecolor="black",
               linestyle="--", label="Authorized investment")
    ]

    leg1 = ax.legend(handles=color_handles, title="Scenarios",
                     loc="center left", bbox_to_anchor=(1.02, 0.62))
    ax.add_artist(leg1)

    ax.legend(handles=style_handles, loc="center left", bbox_to_anchor=(1.02, 0.2), title="Type")

    plt.tight_layout()
    plt.show()


# In[1016]:


result_brut8 = resultV9(result_data2, 'brut2', 'Optimization without constraint')


# In[1017]:


result_corrHC = resultV9(result_data2, 'corr_HC2', 'Peat substitution Limits in Hobby Gardening')


# In[1018]:


result_corrpeat = resultV9(result_data2, 'corr_peat2', 'Peat market limit')


# In[1009]:


result_all_corr = resultV9(result_data2, 'corr_peat_HC2', 'All contraints')


# In[965]:


#ajout des unités


# In[936]:


import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import re
import textwrap
from matplotlib.lines import Line2D
import matplotlib.patheffects as pe  # halo blanc pour séparer les courbes


def resultV10(data, sheet, title_of_graph,
             overlap_eps=1e-6,   # tolérance superposition (sur valeurs normalisées)
             wrap_chars=36       # retour à la ligne dans les encadrés
            ):
    scenario_col = "Scenario"
    indicators = ["Carbon", "EQ", "HH", "EQ (excl. CO2)", "HH (excl. CO2)"]
    inv_tag = "_inv"

    df = pd.read_excel(data, sheet)

    # conversion numérique
    for c in indicators:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # ---------- ordre Excel pour la légende Scenarios ----------
    excel_order = df[scenario_col].tolist()

    def extract_base(name):
        return str(name).replace(inv_tag, "")

    # unique en gardant l’ordre d’apparition Excel
    base_scenarios = list(dict.fromkeys([extract_base(s) for s in excel_order]))

    # ---------- normalisation (sur valeurs brutes) ----------
    vals = df[indicators].copy()
    vmin = vals.min(axis=0)
    vmax = vals.max(axis=0)
    denom = (vmax - vmin).replace(0, np.nan)
    vals_norm = ((vals - vmin) / denom).fillna(0.0)

    # ---------- palette forte (peu de nuances) ----------
    distinct_13 = [
        "#1f77b4",  # bleu
        "#d62728",  # rouge
        "#2ca02c",  # vert
        "#ff7f0e",  # orange
        "#9467bd",  # violet
        "#8c564b",  # brun
        "#17becf",  # cyan
        "#bcbd22",  # olive
        "#e377c2",  # rose
        "#7f7f7f",  # gris
        "#444444",  # anthracite
        "#a65628",  # marron/orangé
        "#c2b280",  # beige foncé
    ]
    color_map = {base: distinct_13[i % len(distinct_13)] for i, base in enumerate(base_scenarios)}

    # ---------- interpolation ----------
    k = 4
    x_base = np.arange(len(indicators))
    x_dense = np.linspace(x_base.min(), x_base.max(), (len(indicators) - 1) * k + 1)

    # marqueurs uniquement sur les traits verticaux
    mark_positions = np.arange(0, len(x_dense), k)

    fig, ax = plt.subplots(figsize=(11, 5))

    # marges : place à gauche pour encadrés + à droite pour légendes
    fig.subplots_adjust(left=0.32, right=0.80)

    # axes verticaux discrets
    for xi in x_base:
        ax.axvline(x=xi, color="gray", linewidth=0.8, alpha=0.35, zorder=0)

    # paramètres visibilité
    lw = 1.25
    a = 1.0
    ms = 6.2
    mew = 1.6

    # =========================
    # Tracé de toutes les courbes
    # =========================
    for i, row in df.iterrows():
        scen = str(row[scenario_col])
        base = extract_base(scen)

        y_base = vals_norm.iloc[i].values
        y_dense = np.interp(x_dense, x_base, y_base)

        is_inv = inv_tag in scen
        marker_style = "x" if is_inv else "o"
        line_style = "--" if is_inv else "-"

        line, = ax.plot(
            x_dense, y_dense,
            linestyle=line_style,
            color=color_map.get(base, "#333333"),
            linewidth=lw,
            alpha=a,
            marker=marker_style,
            markersize=ms,
            markeredgewidth=mew,
            markeredgecolor="black",
            markerfacecolor=color_map.get(base, "#333333") if marker_style == "o" else None,
            markevery=mark_positions,
            zorder=2
        )

        # halo blanc pour séparer les courbes aux croisements
        line.set_path_effects([
            pe.Stroke(linewidth=lw + 2.2, foreground="white", alpha=0.9),
            pe.Normal()
        ])

    # =========================
    # X labels + unités dessous
    # =========================
    units = {
        "Carbon": "kgCO2eq",
        "EQ": "PDF.m2.year",
        "HH": "DALY",
        "EQ (excl. CO2)": "PDF.m2.year",
        "HH (excl. CO2)": "DALY",
    }
    xtlbls = [f"{ind}\n{units.get(ind,'')}" for ind in indicators]
    ax.set_xticks(x_base)
    ax.set_xticklabels(xtlbls)

    # =========================
    # 5 graduations sur chaque ligne verticale (valeurs réelles)
    # =========================
    ticks_norm = np.linspace(0.0, 1.0, 5)   # 0, .25, .5, .75, 1
    tick_halfwidth = 0.06
    label_dx = 0.08
    fs_tick = 7

    for xi, ind in zip(x_base, indicators):
        d = (vmax[ind] - vmin[ind])
        if pd.isna(d) or d == 0:
            continue

        for t in ticks_norm:
            y = float(t)
            val = float(vmin[ind] + t * d)

            # petit trait horizontal (graduation)
            ax.plot([xi - tick_halfwidth, xi + tick_halfwidth],
                    [y, y],
                    color="black", lw=0.5, alpha=0.5, zorder=1)

            # étiquette (valeur réelle)
            ax.text(xi + label_dx, y, f"{val:.2e}",
                    ha="left", va="center",
                    fontsize=fs_tick, color="black", alpha=1,
                    zorder=1)

    # axes / style
    ax.set_ylim(-0.08, 1.08)
    ax.set_yticks([])
    ax.spines["top"].set_visible(False)
    ax.grid(False)
    ax.set_title(title_of_graph)

    # =========================================================
    # Détection des courbes superposées (sur valeurs normalisées)
    # =========================================================
    Y = vals_norm.values
    n = Y.shape[0]

    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            if np.max(np.abs(Y[i] - Y[j])) <= overlap_eps:
                adj[i].append(j)
                adj[j].append(i)

    # composantes connexes = groupes superposés
    seen = [False] * n
    groups = []
    for i in range(n):
        if seen[i]:
            continue
        stack = [i]
        comp = []
        seen[i] = True
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in adj[u]:
                if not seen[v]:
                    seen[v] = True
                    stack.append(v)
        if len(comp) >= 2:
            groups.append(comp)

    # =========================================================
    # Annotations alignées (flèches même longueur, encadrés alignés à droite)
    # =========================================================
    if groups:
        # tri des groupes par position (y au 1er axe)
        def group_y_left(g):
            return float(np.mean([Y[idx][0] for idx in g]))

        groups_sorted = sorted(groups, key=group_y_left, reverse=True)

        # colonne fixe pour les encadrés (alignés côté droit)
        x_box_ax = -0.04      # position des encadrés (axes coords)
        x_target = x_base[0]  # flèche vers le 1er axe vertical

        # positions verticales régulières -> alignement + évite chevauchement
        y_positions = np.linspace(0.85, 0.15, len(groups_sorted))

        for g, y_box in zip(groups_sorted, y_positions):
            scen_list = [str(df.loc[idx, scenario_col]) for idx in g]
            # garder un tri “raisonnable” dans le texte
            scen_list = sorted(scen_list, key=lambda s: excel_order.index(s) if s in excel_order else 10**9)

            text = ", ".join(scen_list)
            if len(text) > wrap_chars:
                text = "\n".join(textwrap.wrap(text, width=wrap_chars, break_long_words=False))

            ax.annotate(
                text,
                xy=(x_target, group_y_left(g)),
                xycoords="data",
                xytext=(x_box_ax, y_box),
                textcoords=ax.transAxes,
                ha="right",   # ✅ alignement côté droit des encadrés
                va="center",
                fontsize=8,
                color="black",
                bbox=dict(boxstyle="round,pad=0.28", fc="white", ec="#222222", lw=1.1, alpha=0.98),
                arrowprops=dict(
                    arrowstyle="-|>",
                    lw=1.8,
                    color="#333333",
                    mutation_scale=18,
                    shrinkA=5,
                    shrinkB=5
                ),
                clip_on=False,
                zorder=10
            )

    # =========================
    # Légendes
    # =========================
    # Scenarios : ordre Excel
    color_handles = [
        Line2D([0], [0], color=color_map[base], lw=2, label=base)
        for base in base_scenarios
    ]

    style_handles = [
        Line2D([0], [0], color="black", lw=lw, marker="o",
               markersize=6, markeredgewidth=1.4, markeredgecolor="black",
               linestyle="-", label="No investment"),
        Line2D([0], [0], color="black", lw=lw, marker="x",
               markersize=6, markeredgewidth=1.6, markeredgecolor="black",
               linestyle="--", label="Authorized investment")
    ]

    leg1 = ax.legend(handles=color_handles, title="Scenarios",
                     loc="center left", bbox_to_anchor=(1.02, 0.62))
    ax.add_artist(leg1)

    ax.legend(handles=style_handles, title="Type",
              loc="center left", bbox_to_anchor=(1.02, 0.2))

    plt.tight_layout()
    plt.show()


# In[937]:


result_brut10 = resultV10(result_data, 'brut2', 'Optimization without constraint')


# In[ ]:


##figure AS


# In[983]:


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

file_path = "fig_AS_ss_inv.xlsx"

# --- Lecture Excel (si l'en-tête n'est pas sur la 1ère ligne, essayer header=1,2,...)
df = pd.read_excel(file_path)

# --- Nettoyage noms de colonnes (espaces, retours, etc.)
df.columns = [str(c).strip() for c in df.columns]

# --- Harmoniser les noms attendus
rename_map = {
    "name": "Name",
    "Sorting rate": "Sorting rate",
    "sorting rate": "Sorting rate",
    "HC": "HC",
    "IC": "IC",
    "AD-cog-comp": "AD-cog-comp",
    "AD-biomet-dig": "AD-biomet-dig",
}
# Appliquer uniquement si la colonne existe
df = df.rename(columns={c: rename_map[c] for c in df.columns if c in rename_map})

# --- Vérif colonnes nécessaires
required = ["Name", "Sorting rate", "HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
missing = [c for c in required if c not in df.columns]
if missing:
    raise ValueError(
        f"Colonnes manquantes: {missing}\n"
        f"Colonnes disponibles: {list(df.columns)}\n"
        "👉 Astuce: si l'en-tête Excel n'est pas sur la première ligne, essaye pd.read_excel(..., header=1) (ou 2,3)."
    )

# --- Nettoyer les valeurs texte
df["Name"] = df["Name"].astype(str).str.strip()
df["Sorting rate"] = df["Sorting rate"].astype(str).str.strip()

# --- Forcer numerique
tech_cols = ["HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
for c in tech_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# --- Ordres (adaptables)
scenario_order = [
    "minCO2", "minEQ", "minHH", "minEQ&HH",
    "minCO2&EQ", "minCO2&HH", "minCO2&EQ&HH"
]
sorting_rates = ["-50%", "0.6", "50%"]  # adapte si tes libellés diffèrent

# Garder seulement les lignes utiles
df = df[df["Name"].isin(scenario_order)].copy()

df["Name"] = pd.Categorical(df["Name"], categories=scenario_order, ordered=True)
df["Sorting rate"] = pd.Categorical(df["Sorting rate"], categories=sorting_rates, ordered=True)
df = df.sort_values(["Name", "Sorting rate"])

# --- Construire la grille complète (au cas où il manque une ligne)
idx = pd.MultiIndex.from_product([scenario_order, sorting_rates], names=["Name", "Sorting rate"])
df = df.set_index(["Name", "Sorting rate"]).reindex(idx).reset_index()
df[tech_cols] = df[tech_cols].fillna(0)

# --- Graphique: 3 barres côte à côte par scénario, empilé 100%
x = np.arange(len(scenario_order))
bar_w = 0.23
offsets = [-bar_w, 0, bar_w]

fig, ax = plt.subplots(figsize=(14, 6))

for i, rate in enumerate(sorting_rates):
    sub = df[df["Sorting rate"] == rate].set_index("Name").loc[scenario_order].reset_index()
    bottom = np.zeros(len(scenario_order))
    for tech in tech_cols:
        vals = sub[tech].to_numpy()
        ax.bar(x + offsets[i], vals, width=bar_w, bottom=bottom, label=tech if i == 0 else "")
        bottom += vals

# --- Mise en forme
ax.set_xticks(x)
ax.set_xticklabels(scenario_order, rotation=45, ha="right")
ax.set_ylim(0, 100)
ax.set_ylabel("Part (%)")
ax.set_title("Sensitivity analysis – Sorting rate impact on technology allocation")
ax.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

# Ajouter une petite légende pour les 3 barres (optionnel)
# (on met juste les positions, et tu peux commenter si tu n'en veux pas)
for i, rate in enumerate(sorting_rates):
    ax.text(x[0] + offsets[i], -5, rate, ha="center", va="top", fontsize=9, transform=ax.transData)

plt.tight_layout()
plt.show()


# In[981]:


df = pd.read_excel('fig_AS_ss_inv.xlsx')


# In[982]:


df


# In[985]:


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

file_path = "fig_AS_ss_inv.xlsx"
df = pd.read_excel(file_path)

# --- Nettoyage colonnes
df.columns = [str(c).strip() for c in df.columns]

# Ici tes colonnes s'appellent "name" (minuscule) et "Sorting rate"
df["name"] = df["name"].astype(str).str.strip()
df["Sorting rate"] = df["Sorting rate"].astype(str).str.strip()

tech_cols = ["HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
for c in tech_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce")

# --- IMPORTANT : gérer les doublons (non-unique multi-index)
# Si tu es sûr qu'il ne devrait pas y avoir de doublons, tu peux aussi faire .mean()
df = (
    df.groupby(["name", "Sorting rate"], as_index=False)[tech_cols]
      .sum(min_count=1)
)

# --- Ordre des scénarios (optionnel)
scenario_order = [
    "minCO2", "minEQ", "minHH", "minEQ&HH",
    "minCO2&EQ", "minCO2&HH", "minCO2&EQ&HH"
]
# On garde uniquement ceux présents
scenario_order = [s for s in scenario_order if s in df["name"].unique()]

sorting_rates = ["-50%", "0.6", "50%"]

df["name"] = pd.Categorical(df["name"], categories=scenario_order, ordered=True)
df["Sorting rate"] = pd.Categorical(df["Sorting rate"], categories=sorting_rates, ordered=True)
df = df.sort_values(["name", "Sorting rate"])

# --- Grille complète (au cas où il manque une ligne)
idx = pd.MultiIndex.from_product([scenario_order, sorting_rates], names=["name", "Sorting rate"])
df = (
    df.set_index(["name", "Sorting rate"])
      .reindex(idx)
      .fillna(0)
      .reset_index()
)

# --- Plot
x = np.arange(len(scenario_order))
bar_w = 0.23
offsets = [-bar_w, 0, bar_w]

fig, ax = plt.subplots(figsize=(14, 6))

for i, rate in enumerate(sorting_rates):
    sub = df[df["Sorting rate"] == rate].set_index("name").loc[scenario_order].reset_index()
    bottom = np.zeros(len(scenario_order))
    for tech in tech_cols:
        vals = sub[tech].to_numpy()
        ax.bar(x + offsets[i], vals, width=bar_w, bottom=bottom, label=tech if i == 0 else "")
        bottom += vals

ax.set_xticks(x)
ax.set_xticklabels(scenario_order, rotation=45, ha="right")
ax.set_ylim(0, 100)
ax.set_ylabel("Part (%)")
ax.set_title("Sensitivity analysis – technology allocation vs sorting rate")
ax.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

plt.tight_layout()
plt.show()


# In[986]:


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

file_path = "fig_AS_ss_inv.xlsx"
df = pd.read_excel(file_path)
df.columns = [str(c).strip() for c in df.columns]

# Colonnes (chez toi: name)
df["name"] = df["name"].astype(str)
df["Sorting rate"] = df["Sorting rate"].astype(str)

# --- Nettoyage fort (enlève espaces multiples, espaces insécables, retours)
def clean_text(s):
    s = s.replace("\u00A0", " ")   # espace insécable
    s = s.replace("\n", " ").replace("\r", " ")
    s = " ".join(s.split())        # compacte espaces
    return s.strip()

df["name"] = df["name"].map(clean_text)
df["Sorting rate"] = df["Sorting rate"].map(clean_text)

# --- Normaliser sorting rate
def norm_rate(x: str) -> str:
    s = x.replace(" ", "").replace(",", ".")
    if s in ["-50%", "-50"]:
        return "-50%"
    if s in ["50%", "+50%", "50", "+50"]:
        return "50%"
    if s in ["0.6", "0.60", "0,6"]:
        return "0.6"
    return s

df["Sorting rate"] = df["Sorting rate"].map(norm_rate)

tech_cols = ["HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
for c in tech_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# --- Agréger les doublons (si doublons accidentels, on somme)
df = df.groupby(["name", "Sorting rate"], as_index=False)[tech_cols].sum()

# --- Ordre
scenario_order = [
    "minCO2", "minEQ", "minHH", "minEQ&HH",
    "minCO2&EQ", "minCO2&HH", "minCO2&EQ&HH"
]
scenario_order = [s for s in scenario_order if s in df["name"].unique()]
sorting_rates = ["-50%", "0.6", "50%"]

# --- Pivot : index=scénarios, colonnes=sorting rate, valeurs=techno
# On obtient un tableau (scenario x rate) pour chaque techno
pivot = {}
for tech in tech_cols:
    pivot[tech] = (df.pivot(index="name", columns="Sorting rate", values=tech)
                     .reindex(index=scenario_order, columns=sorting_rates)
                     .fillna(0))

# --- Plot : 3 barres côte à côte, empilé
x = np.arange(len(scenario_order))
bar_w = 0.23
offsets = [-bar_w, 0, bar_w]

fig, ax = plt.subplots(figsize=(14, 6))

for i, rate in enumerate(sorting_rates):
    bottom = np.zeros(len(scenario_order))
    for tech in tech_cols:
        vals = pivot[tech][rate].to_numpy()
        ax.bar(x + offsets[i], vals, width=bar_w, bottom=bottom, label=tech if i == 0 else "")
        bottom += vals

ax.set_xticks(x)
ax.set_xticklabels(scenario_order, rotation=45, ha="right")
ax.set_ylim(0, 100)
ax.set_ylabel("Part (%)")
ax.set_title("Sensitivity analysis – technology allocation vs sorting rate")
ax.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

plt.tight_layout()
plt.show()


# In[987]:


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

file_path = "fig_AS_ss_inv.xlsx"
df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()

# Nettoyage
df["name"] = df["name"].astype(str).str.strip()
df["Sorting rate"] = df["Sorting rate"].astype(str).str.strip()

tech_cols = ["HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
for c in tech_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# Agréger si doublons
df = df.groupby(["name", "Sorting rate"], as_index=False)[tech_cols].sum()

scenario_order = [
    "minCO2", "minEQ", "minHH", "minEQ&HH",
    "minCO2&EQ", "minCO2&HH", "minCO2&EQ&HH"
]
scenario_order = [s for s in scenario_order if s in df["name"].unique()]

sorting_rates = ["-50%", "0.6", "50%"]

# Pivot
pivot = {}
for tech in tech_cols:
    pivot[tech] = (
        df.pivot(index="name", columns="Sorting rate", values=tech)
          .reindex(index=scenario_order, columns=sorting_rates)
          .fillna(0)
    )

# --- Paramètres graphiques
bar_width = 0.25
group_space = 1.2  # espace entre scénarios

x = np.arange(len(scenario_order)) * group_space

fig, ax = plt.subplots(figsize=(14,6))

offsets = [-bar_width, 0, bar_width]

for i, rate in enumerate(sorting_rates):
    bottom = np.zeros(len(scenario_order))
    for tech in tech_cols:
        values = pivot[tech][rate].values
        ax.bar(x + offsets[i], values, bar_width,
               bottom=bottom,
               label=tech if i == 0 else "")
        bottom += values

# Axe X principal (scénarios)
ax.set_xticks(x)
ax.set_xticklabels(scenario_order, rotation=45, ha="right")

# Ajouter sous-labels -50 / N / +50
for i, rate in enumerate(sorting_rates):
    for j in range(len(scenario_order)):
        ax.text(x[j] + offsets[i], -5, rate,
                ha='center', va='top', fontsize=9)

ax.set_ylim(0, 100)
ax.set_ylabel("Part (%)")
ax.set_title("Sensitivity analysis – Technology allocation vs sorting rate")

ax.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

plt.tight_layout()
plt.show()


# In[988]:


import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

file_path = "fig_AS_ss_inv.xlsx"
df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()

# Nettoyage
df["name"] = df["name"].astype(str).str.strip()
df["Sorting rate"] = df["Sorting rate"].astype(str).str.strip()

tech_cols = ["HC", "IC", "AD-cog-comp", "AD-biomet-dig"]
for c in tech_cols:
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

# Agréger si doublons
df = df.groupby(["name", "Sorting rate"], as_index=False)[tech_cols].sum()

scenario_order = [
    "minCO2", "minEQ", "minHH", "minEQ&HH",
    "minCO2&EQ", "minCO2&HH", "minCO2&EQ&HH"
]
scenario_order = [s for s in scenario_order if s in df["name"].unique()]

sorting_rates = ["-50%", "0.6", "50%"]

# Pivot
pivot = {}
for tech in tech_cols:
    pivot[tech] = (
        df.pivot(index="name", columns="Sorting rate", values=tech)
          .reindex(index=scenario_order, columns=sorting_rates)
          .fillna(0)
    )

# --- Paramètres graphiques
bar_width = 0.25
group_space = 1.2  # espace entre scénarios

x = np.arange(len(scenario_order)) * group_space

fig, ax = plt.subplots(figsize=(14,6))

offsets = [-bar_width, 0, bar_width]

for i, rate in enumerate(sorting_rates):
    bottom = np.zeros(len(scenario_order))
    for tech in tech_cols:
        values = pivot[tech][rate].values
        ax.bar(x + offsets[i], values, bar_width,
               bottom=bottom,
               label=tech if i == 0 else "")
        bottom += values

# Axe X principal (scénarios)
ax.set_xticks(x)
ax.set_xticklabels(scenario_order, rotation=45, ha="right")

# Ajouter sous-labels -50 / N / +50
for i, rate in enumerate(sorting_rates):
    for j in range(len(scenario_order)):
        ax.text(x[j] + offsets[i], -5, rate,
                ha='center', va='top', fontsize=9)

ax.set_ylim(0, 100)
ax.set_ylabel("Part (%)")
ax.set_title("Sensitivity analysis – Technology allocation vs sorting rate")

ax.legend(title="Technology", bbox_to_anchor=(1.02, 1), loc="upper left")

plt.tight_layout()
plt.show()


# In[ ]:




