#!/usr/bin/env python
# coding: utf-8

# In[18]:


import numpy as np
import pandas as pd
from collections import defaultdict
import pulp


# In[19]:


I_brut = 'impact_brut.xlsx'
ARC_reel = 'ARC_reel.xlsx'
tech_ARC_reel = 'tech_ARC_reel.xlsx'
I_hc_correct = 'impact_HC_correct.xlsx'
I_peat_corr = 'impact_peat_corr.xlsx'
I_peat_HC_corr = 'impact_peat_HC_corr.xlsx'
ARC_reel_ss_inv = 'ARC_reel_sans_inv.xlsx'
tech_reel_ss_inv = 'tech_ARC_reel_sans_inv.xlsx'
peat_corr2 = 'sim_peat_corr2.xlsx'
result_data = 'result_norm_vf2.xlsx'


# In[6]:


class opt:  
    def __init__(self, municipalities, technologies, method, sort, ind, ind1, ind2, ind3, ind4, W, W1, W2, W3, W4,                M_peat):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.ind3 = ind3
        self.ind4 = ind4
        self.W = W
        self.W1 = W1
        self.W2 = W2
        self.W3 = W3
        self.W4 = W4
        self.M_peat = M_peat
        

        
        #current scenario's impact (incineration)

        
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'coproduct')
            self.coprod_list = np.array(coprod_list['Co-product'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            self.unit3 = imp[self.ind3][0]
            self.unit4 = imp[self.ind4][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            self.tr_w3 = imp[self.ind3][1]
            self.tr_w4 = imp[self.ind4][1]
            
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            self.col_w3 = imp[self.ind3][2]
            self.col_w4 = imp[self.ind4][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            self.I_T3 = tech[self.ind3]
            self.I_T4 = tech[self.ind4]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            self.I_inc3 = self.I_T3['Inc']
            self.I_inc4 = self.I_T4['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            self.I_tr_cop3 = tr_cop[self.ind3]
            self.I_tr_cop4 = tr_cop[self.ind4]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            self.I_av3 = av[self.ind3]
            self.I_av4 = av[self.ind4]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
            self.I_u_cop3 = u_cop[self.ind3]
            self.I_u_cop4 = u_cop[self.ind4]
            
        self.N = self.tot_mass * self.I_inc
        self.N1 = self.tot_mass * self.I_inc1
        self.N2 = self.tot_mass * self.I_inc2
        self.N3 = self.tot_mass * self.I_inc3
        self.N4 = self.tot_mass * self.I_inc4
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"y_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        trans_I3 = {}
        trans_I4 = {}
        
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans_I3[(M,S)] = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans_I4[(M,S)] = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
        
        #trans_I1_results = pd.DataFrame(list(trans_I1.items()), columns=["trajet", "Values"])
        #trans_I1_results.to_excel("trans_EQ.xlsx", index=False)
        
        #trans_I2_results = pd.DataFrame(list(trans_I2.items()), columns=["trajet", "Values"])
        #trans_I2_results.to_excel("trans_HH.xlsx", index=False)
        
        #trans_I_results = pd.DataFrame(list(trans_I.items()), columns=["trajet", "Values"])
        #trans_I_results.to_excel("trans_C.xlsx", index=False)
        

                        
        #impact of transport of coproduct
        trans_I_cp1 = {}
        trans_I_cp2 = {}
        trans_I_cp = {}
        trans_I_cp3 = {}
        trans_I_cp4 = {}
        for S in self.S_list:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    trans_I_cp1[(S,cp,mark)] = self.I_tr_cop1[cp] * self.dist[mark][S]
                    trans_I_cp2[(S,cp,mark)] = self.I_tr_cop2[cp] * self.dist[mark][S]
                    trans_I_cp[(S,cp,mark)] = self.I_tr_cop[cp] * self.dist[mark][S]
                    trans_I_cp3[(S,cp,mark)] = self.I_tr_cop3[cp] * self.dist[mark][S]
                    trans_I_cp4[(S,cp,mark)] = self.I_tr_cop4[cp] * self.dist[mark][S]
        
        #trans_I_cp1_results = pd.DataFrame(list(trans_I_cp1.items()), columns=["trajet", "Values"])
        #trans_I_cp1_results.to_excel("cp_EQ.xlsx", index=False)
        
        #trans_I_cp2_results = pd.DataFrame(list(trans_I_cp2.items()), columns=["trajet", "Values"])
        #trans_I_cp2_results.to_excel("cp_HH.xlsx", index=False)
        
        #trans_I_cp_results = pd.DataFrame(list(trans_I_cp.items()), columns=["trajet", "Values"])
        #trans_I_cp_results.to_excel("cp_C.xlsx", index=False)
        
        
        #impact of substitution/avoided impact of peat
        self.sub_I = {}
        self.sub_I1 = {}
        self.sub_I2 = {}
        self.sub_I3 = {}
        self.sub_I4 = {}
    
        for T in self.idtech:  # technologies (HC, IC, AD-a, AD-b)
            #for cp in self.coprod_list:  # coproduits : Comp, Dig, Heat, Elect
           #     for conv_p in self.c_p_list:  # produits conventionnels : Fertilizer, Heat, Ele
            self.sub_I[T] = self.avoid.loc['Peat', T]*(self.I_av['Peat']+ (0.7318*1))
            self.sub_I1[T] = self.avoid.loc['Peat', T]*(self.I_av1['Peat']+ (0.7318*0.702543))
            self.sub_I2[T] = self.avoid.loc['Peat', T]*(self.I_av2['Peat'] +(0.7318*7.1373e-6))
            self.sub_I3[T] = self.avoid.loc['Peat', T]*(self.I_av3['Peat'] +(0.7318*0.1685))
            self.sub_I4[T] = self.avoid.loc['Peat', T]*(self.I_av4['Peat'] +(0.7318*0))

        
        
        
        #print("=== Substitution peat coefficients (par tonne de déchets traités) ===")
        #for T in self.idtech:
        #    print(T, "avoid=", self.avoid.at[T,"Peat"] if ("Peat" in self.avoid.columns and T in self.avoid.index) else None,
        #  "sub_I=", self.sub_I[T], self.sub_I1[T], self.sub_I2[T], self.sub_I3[T], self.sub_I4[T])
        
        #print("avoid index (5):", list(self.avoid.index)[:5])
        #print("avoid columns:", list(self.avoid.columns))


        
        
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
        # --- Plafonds de marché (dans les unités des indicateurs) ---
        #self.X_CO2  = 0# 10000000 #6.45E+04 # kgCO2e max évités (ton ind "Carbon")
        #self.X_EQ_ssC  =0# 1e10 # 11369.41689
   # PDF max évités (ton ind1)
        #self.X_HH_ssC = 0# 1e10 #3.82E-03# DALY max évités (ton ind2)
        #self.X_EQ = 0
        #self.X_HH = 0
        
        
        self.X_CO2 = self.M_peat*(self.I_av['Peat']+ (0.7318*1))
        self.X_EQ = self.M_peat*(self.I_av1['Peat']+ (0.7318*0.702543))
        self.X_HH = self.M_peat* (self.I_av2['Peat'] +(0.7318*7.1373e-6))
        self.X_EQssC = self.M_peat* (self.I_av3['Peat'] +(0.7318*0.1685))
        self.X_HHssC = self.M_peat* (self.I_av4['Peat'] +(0.7318*0))
        

# évitements bruts
        
        self.A_CO2_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        self.A_EQ_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I1[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

        self.A_HH_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I2[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        
        self.A_EQssC_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I3[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)
        
        self.A_HHssC_raw = pulp.lpSum(x[(M, O, T, S)] * self.sub_I4[T]
    for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list)

# variables capées
        self.A_CO2_cap  = pulp.LpVariable("Peat_avoided_CO2_cap",  lowBound=0)
        self.A_EQ_cap = pulp.LpVariable("Peat_avoided_EQ_cap", lowBound=0)
        self.A_HH_cap = pulp.LpVariable("Peat_avoided_HH_cap", lowBound=0)
        self.A_EQssC_cap  = pulp.LpVariable("Peat_avoided_EQssC_cap",  lowBound=0)
        self.A_HHssC_cap = pulp.LpVariable("Peat_avoided_HHssC_cap", lowBound=0)

# linéarisation du min(., X)
        model += self.A_CO2_cap  <= self.A_CO2_raw
        model += self.A_CO2_cap  <= self.X_CO2


        
        model += self.A_EQ_cap <= self.A_EQ_raw
        model += self.A_EQ_cap <=self.X_EQ
        
        model += self.A_HH_cap <= self.A_HH_raw
        model += self.A_HH_cap <=self.X_HH
        
        model += self.A_EQssC_cap  <= self.A_EQssC_raw
        model += self.A_EQssC_cap  <= self.X_EQssC

        model += self.A_HHssC_cap <= self.A_HHssC_raw
        model += self.A_HHssC_cap <=self.X_HHssC
        
    



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T]))+\
                                               ((self.W3/self.N3)*((trans_I3[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T3[T]))+\
                                               ((self.W4/self.N4)*((trans_I4[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T4[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        + pulp.lpSum(z[(T, S, cp, mark)] * (((self.W1/self.N1)*trans_I_cp1[(S,cp,mark)])+\
                                            ((self.W2/self.N2)*trans_I_cp2[(S,cp,mark)])+\
                                           ((self.W/self.N)*trans_I_cp[(S,cp,mark)])+\
                                           ((self.W3/self.N3)*trans_I_cp3[(S,cp,mark)])+\
                                           ((self.W4/self.N4)*trans_I_cp4[(S,cp,mark)]))\
                     for T in self.idtech for cp in self.coprod_list for mark in self.market_list for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))+\
                          ((self.W3/self.N3)*(self.I_inc3))+((self.W4/self.N4)*(self.I_inc4))))\
            - ((self.W/self.N)*self.A_CO2_cap) - ((self.W1/self.N1) * self.A_EQ_cap)\
        - ((self.W2/self.N2) * self.A_HH_cap) - ((self.W3/self.N3)*self.A_EQssC_cap) - ((self.W4/self.N4)*self.A_HHssC_cap)

        model.solve()
        
        #print("=== Market cap check ===")
        #print("CO2  raw:", pulp.value(self.A_CO2_raw), "cap:", pulp.value(self.A_CO2_cap), "X:", self.X_CO2)
        #print("EQssC  raw:", pulp.value(self.A_EQssC_raw), "cap:", pulp.value(self.A_EQssC_cap), "X:", self.X_EQssC)
        #print("HHssC raw:", pulp.value(self.A_HHssC_raw), "cap:", pulp.value(self.A_HHssC_cap), "X:", self.X_HHssC)
        #print("EQ raw:", pulp.value(self.A_EQ_raw), "cap:", pulp.value(self.A_EQ_cap), "X:", self.X_EQ)
        #print("HH raw:", pulp.value(self.A_HH_raw), "cap:", pulp.value(self.A_HH_cap), "X:", self.X_HH)

        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        total3 = 0
        total4 = 0
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans3 = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans4 = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0
                trans3 = 0
                trans4 = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T] )# - self.sub_I1[T])
            total2 += val * (trans2 + self.I_T2[T] ) #- self.sub_I2[T])
            total += val * (trans + self.I_T[T])
            total3 += val * (trans3 + self.I_T3[T])
            total4 += val * (trans4 + self.I_T4[T])# - self.sub_I[T])
        


    

        # ---------- Transport des coproduits ----------
        for (T, S, cp, mark), val in self.z_opt.items():
            if self.scale[T] == 'Centralized':
                total1 += val * self.I_tr_cop1[cp] * self.dist[mark][S]
                total2 += val * self.I_tr_cop2[cp] * self.dist[mark][S]
                total += val * self.I_tr_cop[cp] * self.dist[mark][S]
                total3 += val * self.I_tr_cop3[cp] * self.dist[mark][S]
                total4 += val * self.I_tr_cop4[cp] * self.dist[mark][S]
            else:
                total1 += 0
                total2 += 0
                total += 0
                total3 += 0
                total4 += 0


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        total3 += self.unsort * self.I_inc3
        total4 += self.unsort * self.I_inc4
        
            # ---------- Crédit tourbe (capé par le marché) ----------
        A1_raw = sum(val * self.sub_I1[T] for (M, O, T, S), val in self.x_opt.items())
        A2_raw = sum(val * self.sub_I2[T] for (M, O, T, S), val in self.x_opt.items())
        A0_raw = sum(val * self.sub_I[T]  for (M, O, T, S), val in self.x_opt.items())
        A3_raw = sum(val * self.sub_I3[T]  for (M, O, T, S), val in self.x_opt.items())
        A4_raw = sum(val * self.sub_I4[T]  for (M, O, T, S), val in self.x_opt.items())

        A1_cap = min(A1_raw, self.X_EQ)
        A2_cap = min(A2_raw, self.X_HH)
        A0_cap = min(A0_raw, self.X_CO2)
        A3_cap = min(A3_raw, self.X_EQssC)
        A4_cap = min(A4_raw, self.X_HHssC)
        

        total1 -= A1_cap
        total2 -= A2_cap
        total  -= A0_cap
        total3  -= A3_cap
        total4  -= A4_cap
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        print(f"Impact total {self.ind3} : {total3:.2f} {self.unit3}")
        print(f"Impact total {self.ind4} : {total4:.2f} {self.unit4}")
        
        return {self.ind : total, self.ind1: total1, self.ind2 :total2,                self.ind3 : total3 , self.ind4 : total4}
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table
    
    def get_shares_tech_site(self, decimals=2, tol=1e-9, only_feasible=True):
        """
        Retourne les % de flux par (tech, site) sur le total trié traité.
        """
        ts = {}
        total = 0.0
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            if only_feasible and self.poss_tech[T][S] == 0:
                continue
            total += float(v)
            ts[(T, S)] = ts.get((T, S), 0.0) + float(v)

        if total <= tol:
            return {}

        return {(T, S): round(100.0 * f / total, decimals) for (T, S), f in ts.items()}

        
        
        


# In[50]:


sim1 = opt(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 1 , 0, 0, 0, 0, 0)


# In[51]:


sim1.optimize()


# In[9]:


sim1.flows_techno()


# In[10]:


sim1.impact_tot()


# In[12]:


sim7 = opt(ARC_reel, tech_ARC_reel, I_brut, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 0)


# In[13]:


sim7.optimize()


# In[14]:


sim7.impact_tot()


# In[16]:


simHC = opt(ARC_reel, tech_ARC_reel, I_hc_correct, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 0)


# In[17]:


simHC.optimize()


# In[18]:


simHC.impact_tot()


# In[20]:


simpeat = opt(ARC_reel, tech_ARC_reel, I_peat_corr, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 86405.3218)


# In[21]:


simpeat.optimize()


# In[22]:


simpeat.impact_tot()


# In[25]:


simpeatHC = opt(ARC_reel, tech_ARC_reel, I_peat_HC_corr, 0.6,'Carbon', 'EQ', 'HH', 'EQssC', 'HHssC', 0.5 , 0,0, 0.5, 0, 86405.3218)


# In[27]:


simpeatHC.optimize()


# In[28]:


simpeatHC.impact_tot()


# In[23]:




def run_batch_from_excel(scenarios_xlsx,
    municipalities,
    technologies,
    method, sort,
    out_xlsx="results_batch.xlsx",
    sheet="scenarios"):
    
    scen = pd.read_excel(scenarios_xlsx, sheet_name=sheet)
    
    sort = sort

    rows = []
    for _, r in scen.iterrows():
        name = str(r["name"])
        #sort = float(r["sort"])
        M_peat = float(r["M_peat"])

        # poids (W) : adapte si tes colonnes sont nommées autrement
        W  = float(r["W"])
        W1 = float(r["W1"])
        W2 = float(r["W2"])
        W3 = float(r["W3"])
        W4 = float(r["W4"])

        m = opt(
            municipalities=municipalities,
            technologies=technologies,
            method=method,
            sort=sort,
            ind="Carbon",
            ind1="EQ",
            ind2="HH",
            ind3="EQssC",
            ind4="HHssC",
            W=W, W1=W1, W2=W2, W3=W3, W4=W4,
            M_peat=M_peat
        )
        m.optimize()

        impacts = m.impact_tot()
        shares = m.get_shares_tech_site()

        rows.append({
            "name": name,
            "sort": sort,
            "impacts": impacts,
            "shares": shares,
            "idtech": list(m.idtech),
            "S_list": list(m.S_list),
            "poss_tech": m.poss_tech,
        })

    # construire tableau résultat au format techno/site (2 lignes d'en-têtes)
    df = build_output_table(rows)
    export_two_header_excel(df, out_xlsx)
    return df


# In[24]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

IMPACT_ORDER = ["Carbon", "EQ", "HH", "EQssC", "HHssC"]

def build_output_table(rows):
    # on prend la structure techno/site depuis le 1er modèle (identique pour tous scénarios normalement)
    idtech = rows[0]["idtech"]
    S_list = rows[0]["S_list"]
    poss_tech = rows[0]["poss_tech"]

    # colonnes techno/site faisables
    tech_site_cols = []
    for T in idtech:
        for S in S_list:
            try:
                feasible = float(poss_tech[T][S]) != 0
            except Exception:
                feasible = True
            if feasible:
                tech_site_cols.append((T, S))

    columns = [("", "name"), ("", "sort")] + tech_site_cols + [("", c) for c in IMPACT_ORDER]
    cols = pd.MultiIndex.from_tuples(columns)

    data = []
    for r in rows:
        row = {("", "name"): r["name"], ("", "sort"): r["sort"]}

        # % flux
        for (T, S) in tech_site_cols:
            row[(T, S)] = r["shares"].get((T, S), None)

        # impacts
        for c in IMPACT_ORDER:
            row[("", c)] = r["impacts"].get(c, None)

        data.append(row)

    return pd.DataFrame(data, columns=cols)

def export_two_header_excel(df, out_xlsx):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="results", index=True)

    wb = load_workbook(out_xlsx)
    ws = wb["results"]

    # aligner en-têtes
    for cell in ws[1] + ws[2]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # fusionner techno (ligne 1) sur plusieurs sites
    max_col = ws.max_column
    cur = ws.cell(row=1, column=1).value
    start = 1
    for col in range(2, max_col + 2):
        val = ws.cell(row=1, column=col).value if col <= max_col else None
        if val != cur:
            if cur not in (None, "") and (col - start) > 1:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col-1)
            start = col
            cur = val

    # format % sur colonnes techno/site (top != "" et sub != "")
    for col in range(1, max_col + 1):
        top = ws.cell(row=1, column=col).value
        sub = ws.cell(row=2, column=col).value
        if top not in (None, "") and sub not in (None, ""):
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    # format scientifique sur impacts (top=="" et sub in IMPACT_ORDER)
    for col in range(1, max_col + 1):
        top = ws.cell(row=1, column=col).value
        sub = ws.cell(row=2, column=col).value
        if (top in (None, "")) and (sub in IMPACT_ORDER):
            for r in range(3, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00E+00"

    wb.save(out_xlsx)


# In[9]:


peat_corr_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_corr, sort =0.6,
    out_xlsx="peat_corr_inv_result_vf2.xlsx", sheet = 'Feuil1')


# In[38]:


peat_corr_ss_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_corr, sort=0.6,
    out_xlsx="peat_corr_ss_inv_result_vf2.xlsx", sheet = 'Feuil1'
)


# In[39]:


peat_HC_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr, sort = 0.6,
    out_xlsx="peat_HC_inv_result_vf2.xlsx", sheet = 'Feuil1'
)


# In[40]:


peat_HC_ss_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr, sort = 0.6,
    out_xlsx="peat_HC_ss_inv_result_vf2.xlsx", sheet = 'Feuil1'
)


# In[3]:


import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import re
from matplotlib.lines import Line2D
import matplotlib.patheffects as pe
import textwrap

def figure(data, sheet, title_of_graph,
             overlap_eps=1e-6,   # tolérance: 0.0 strict, 1e-6/1e-4 si arrondis
             wrap_chars=36,      # largeur max avant retour ligne dans l'encadré
             min_gap_axes=0.08   # espacement vertical entre encadrés (coords axes)
            ):

    scenario_col = "Scenario"
    indicators = ["Carbon", "EQ", "HH", "EQ (excl. CO2)", "HH (excl. CO2)"]
    inv_tag = "_inv"

    df = pd.read_excel(data, sheet)

    # conversion numérique
    for c in indicators:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # --- tri naturel (même si pas de chiffres dans les noms)
    def extract_base(name):
        return str(name).replace(inv_tag, "")

    def scen_sort_key(name):
        name = str(name)
        base = extract_base(name)
        m = re.search(r"\d+", base)
        num = int(m.group()) if m else 10**9
        is_inv = inv_tag in name
        return (num, is_inv, base, name)

    df = df.sort_values(
        by=scenario_col,
        key=lambda col: col.map(scen_sort_key)
    ).reset_index(drop=True)

    # --- normalisation APRES tri
    vals = df[indicators].copy()
    vmin = vals.min(axis=0)
    vmax = vals.max(axis=0)
    denom = (vmax - vmin).replace(0, np.nan)
    vals_norm = ((vals - vmin) / denom).fillna(0.0)

    # --- liste des scénarios "base" pour couleurs (mais ici tes scénarios ne sont pas numérotés,
    # donc base == scénarios eux-mêmes si pas de _inv)
    base_scenarios =  [
    "minCO2",
    "minEQ",
    "minHH",
    "minEQ&HH",
    "minCO2&EQ",
    "minCO2&HH",
    "minCO2&EQ&HH"
]


    # --- palette forte
    distinct_13 = [
        "#1f77b4",  # bleu
        "#d62728",  # rouge
        "#2ca02c",  # vert
        "#ff7f0e",  # orange
        "#9467bd",  # violet
        "#8c564b",  # brun
        "#17becf",  # cyan
        "#bcbd22",  # olive
        "#e377c2",  # rose
        "#7f7f7f",  # gris
        "#444444",  # anthracite (plus doux que noir)
        "#a65628",  # marron/orangé
        "#c2b280",  # beige foncé
    ]
    color_map = {base: distinct_13[i % len(distinct_13)] for i, base in enumerate(base_scenarios)}

    # --- interpolation
    k = 4
    x_base = np.arange(len(indicators))
    x_dense = np.linspace(x_base.min(), x_base.max(), (len(indicators) - 1) * k + 1)
    mark_positions = np.arange(0, len(x_dense), k)  # marqueurs uniquement sur les axes

    fig, ax = plt.subplots(figsize=(11, 5))

    for xi in x_base:
        ax.axvline(x=xi, color="gray", linewidth=0.8, alpha=0.35, zorder=0)

    # paramètres tracé
    lw = 1.25
    a = 1.0
    ms = 6.2
    mew = 1.6

    # ==============================
    # 1) Tracé de toutes les courbes
    # ==============================
    for i, row in df.iterrows():
        scen = str(row[scenario_col])
        base = extract_base(scen)

        y_base = vals_norm.iloc[i].values
        y_dense = np.interp(x_dense, x_base, y_base)

        is_inv = inv_tag in scen
        marker_style = "x" if is_inv else "o"
        line_style = "--" if is_inv else "-"

        line, = ax.plot(
            x_dense, y_dense,
            linestyle=line_style,
            color=color_map[base],
            linewidth=lw,
            alpha=a,
            marker=marker_style,
            markersize=ms,
            markeredgewidth=mew,
            markeredgecolor="black",
            markerfacecolor=color_map[base] if marker_style == "o" else None,
            markevery=mark_positions,
            zorder=2
        )

        line.set_path_effects([
            pe.Stroke(linewidth=lw + 2.2, foreground="white", alpha=0.9),
            pe.Normal()
        ])

    # ==============================================
    # 2) Détection des courbes superposées
    #    On compare les vecteurs normalisés (5 valeurs)
    # ==============================================
    Y = vals_norm.values
    n = Y.shape[0]

    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            if np.max(np.abs(Y[i] - Y[j])) <= overlap_eps:
                adj[i].append(j)
                adj[j].append(i)

    # composantes connexes = groupes superposés
    seen = [False] * n
    groups = []
    for i in range(n):
        if seen[i]:
            continue
        stack = [i]
        comp = []
        seen[i] = True
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in adj[u]:
                if not seen[v]:
                    seen[v] = True
                    stack.append(v)
        if len(comp) >= 2:
            groups.append(comp)

    # =========================================================
    # 3) Annotations avec flèche pour les groupes superposés
    #    -> on pointe vers la zone de gauche (x=0)
    #    -> encadrés empilés pour ne pas se chevaucher
    # =========================================================
    # =========================================================
# 3) Annotations alignées (flèches même longueur)
# =========================================================

    if groups:

        fig.subplots_adjust(left=0.30)

    # Position horizontale fixe des encadrés (même colonne)
        x_box_ax = -0.08   # colonne fixe (coord axes)
        x_target = x_base[0]  # on pointe vers le premier axe vertical

    # Trier du haut vers le bas
        def group_y_left(g):
            return float(np.mean([Y[idx][0] for idx in g]))

        groups_sorted = sorted(groups, key=group_y_left, reverse=True)

    # Espacement vertical régulier
        n_groups = len(groups_sorted)
        y_positions = np.linspace(0.85, 0.15, n_groups)

        for g, y_box in zip(groups_sorted, y_positions):

            scen_list = [str(df.loc[idx, scenario_col]) for idx in g]
            scen_list = sorted(scen_list, key=scen_sort_key)

            text = ", ".join(scen_list)
            text = "\n".join(textwrap.wrap(text, width=36, break_long_words=False))

            ax.annotate(
            text,
            xy=(x_target, group_y_left(g)),
            xycoords="data",
            xytext=(x_box_ax, y_box),
            textcoords=ax.transAxes,
            ha="right",   # ← alignement côté droit
            va="center",
            fontsize=8,
            color="#333333",
            bbox=dict(
                boxstyle="round,pad=0.28",
                fc="white",
                ec="#222222",
                lw=1.1,
                alpha=0.98
            ),
            arrowprops=dict(
                arrowstyle="-|>",
                lw=1.8,
                color="#333333",
                mutation_scale=18,
                shrinkA=5,
                shrinkB=5
            ),
            clip_on=False,
            zorder=10)


    # --- valeurs affichées (haut=max, bas=min) comme ton code
    for xi, ind in zip(x_base, indicators):
        ax.text(xi - 0.07, 1.02, f"{vmax[ind]:.2e}",
                ha="right", va="bottom", fontsize=8, alpha=0.85)
        ax.text(xi - 0.07, -0.02, f"{vmin[ind]:.2e}",
                ha="right", va="top", fontsize=8, alpha=0.85)

    ax.set_xticks(x_base)
    ax.set_xticklabels(indicators)
    ax.set_ylim(-0.1, 1.1)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.grid(False)
    ax.set_title(title_of_graph)

    # --- légendes (comme avant)
    color_handles = [
        Line2D([0], [0], color=color_map[base], lw=2, label=base)
        for base in bases_presentes
    ]

    style_handles = [
        Line2D([0], [0], color="black", lw=lw, marker="o",
               markersize=6, markeredgewidth=1.4, markeredgecolor="black",
               linestyle="-", label="No investment"),
        Line2D([0], [0], color="black", lw=lw, marker="x",
               markersize=6, markeredgewidth=1.6, markeredgecolor="black",
               linestyle="--", label="Authorized investment")
    ]

    leg1 = ax.legend(handles=color_handles, title="Scenarios",
                     loc="center left", bbox_to_anchor=(1.02, 0.62))
    ax.add_artist(leg1)

    ax.legend(handles=style_handles, loc="center left", bbox_to_anchor=(1.02, 0.2), title="Type")

    plt.tight_layout()
    plt.show()


# In[111]:


result_corrpeat = figure(result_data, 'corr_peat2', 'Peat market limit')


# In[44]:


result_all_corr = figure(result_data, 'corr_peat_HC2', 'All contraints')


# In[45]:


ASpos_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr, sort = 0.9,
    out_xlsx="ASpos_inv_vf2.xlsx", sheet = 'Feuil1'
)


# In[46]:


ASpos = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr, sort = 0.9,
    out_xlsx="ASpos_vf2.xlsx", sheet = 'Feuil1'
)


# In[47]:


ASneg_inv = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel,
    technologies=tech_ARC_reel,
    method=I_peat_HC_corr, sort = 0.3,
    out_xlsx="ASneg_inv_vf2.xlsx", sheet = 'Feuil1'
)
print(df.head())


# In[48]:


ASneg = run_batch_from_excel(scenarios_xlsx=peat_corr2,
    municipalities= ARC_reel_ss_inv,
    technologies=tech_reel_ss_inv,
    method=I_peat_HC_corr, sort = 0.3,
    out_xlsx="ASneg_vf2.xlsx", sheet = 'Feuil1'
)


# In[10]:


class opt_corr_peat:  
    def __init__(self, municipalities, technologies, method, sort, ind, ind1, ind2, ind3, ind4, W, W1, W2, W3, W4,                M_hobby, M_pro):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.ind3 = ind3
        self.ind4 = ind4
        self.W = W
        self.W1 = W1
        self.W2 = W2
        self.W3 = W3
        self.W4 = W4
        self.M_hobby = M_hobby
        self.M_pro = M_pro
        

        
        #current scenario's impact (incineration)

        
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'subs_ratio')
            self.coprod_list = np.array(coprod_list['Service'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            self.unit3 = imp[self.ind3][0]
            self.unit4 = imp[self.ind4][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            self.tr_w3 = imp[self.ind3][1]
            self.tr_w4 = imp[self.ind4][1]
            
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            self.col_w3 = imp[self.ind3][2]
            self.col_w4 = imp[self.ind4][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            self.I_T3 = tech[self.ind3]
            self.I_T4 = tech[self.ind4]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            self.I_inc3 = self.I_T3['Inc']
            self.I_inc4 = self.I_T4['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            self.I_tr_cop3 = tr_cop[self.ind3]
            self.I_tr_cop4 = tr_cop[self.ind4]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            self.I_av3 = av[self.ind3]
            self.I_av4 = av[self.ind4]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
            self.I_u_cop3 = u_cop[self.ind3]
            self.I_u_cop4 = u_cop[self.ind4]
            
        self.N = self.tot_mass * self.I_inc
        self.N1 = self.tot_mass * self.I_inc1
        self.N2 = self.tot_mass * self.I_inc2
        self.N3 = self.tot_mass * self.I_inc3
        self.N4 = self.tot_mass * self.I_inc4
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"z_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        trans_I3 = {}
        trans_I4 = {}
        
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans_I3[(M,S)] = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans_I4[(M,S)] = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
        
  
        

                        
        #impact of transport of coproduct


        
        
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
        # --- Plafonds de marché (dans les unités des indicateurs) ---

        
        
        self.X_CO2_hobby = self.M_hobby*(self.I_av['Peat']+ (0.7318*1))
        self.X_EQ_hobby = self.M_hobby*(self.I_av1['Peat']+ (0.7318*0.702543))
        self.X_HH_hobby = self.M_hobby* (self.I_av2['Peat'] +(0.7318*7.1373e-6))
        self.X_EQssC_hobby = self.M_hobby* (self.I_av3['Peat'] +(0.7318*0.1685))
        self.X_HHssC_hobby = self.M_hobby* (self.I_av4['Peat'] +(0.7318*0))
        
        self.X_CO2_pro = self.M_pro*(self.I_av['Peat']+ (0.7318*1))
        self.X_EQ_pro = self.M_pro*(self.I_av1['Peat']+ (0.7318*0.702543))
        self.X_HH_pro = self.M_pro* (self.I_av2['Peat'] +(0.7318*7.1373e-6))
        self.X_EQssC_pro = self.M_pro* (self.I_av3['Peat'] +(0.7318*0.1685))
        self.X_HHssC_pro = self.M_pro* (self.I_av4['Peat'] +(0.7318*0))
        

# évitements bruts
#hobby market
        
        self.A_CO2_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av['Peat']+ (0.7318*1))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_EQ_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av1['Peat']+ (0.7318*0.702543))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)

        self.A_HH_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av2['Peat']+ (0.7318*7.1373e-6))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_EQssC_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av3['Peat']+ (0.7318*0.1685))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_HHssC_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av4['Peat']+ (0.7318*0))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        
   #professionnal market     
        
        self.A_CO2_raw_pro = pulp.lpSum(z[(T, S,cp, 'Peat_pro_mark')] * (self.I_av['Peat']+ (0.7318*1))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_EQ_raw_pro = pulp.lpSum(z[(T, S,cp, 'Peat_pro_mark')] * (self.I_av1['Peat']+ (0.7318*0.702543))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)

        self.A_HH_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av2['Peat']+ (0.7318*7.1373e-6))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_EQssC_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av3['Peat']+ (0.7318*0.1685))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_HHssC_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av4['Peat']+ (0.7318*0))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)

# variables capées
        self.A_CO2_cap_hobby  = pulp.LpVariable("Peat_avoided_CO2_cap_hobby",  lowBound=0)
        self.A_EQ_cap_hobby = pulp.LpVariable("Peat_avoided_EQ_cap_hobby", lowBound=0)
        self.A_HH_cap_hobby = pulp.LpVariable("Peat_avoided_HH_cap_hobby", lowBound=0)
        self.A_EQssC_cap_hobby  = pulp.LpVariable("Peat_avoided_EQssC_cap_hobby",  lowBound=0)
        self.A_HHssC_cap_hobby = pulp.LpVariable("Peat_avoided_HHssC_cap_hobby", lowBound=0)
        
        self.A_CO2_cap_pro  = pulp.LpVariable("Peat_avoided_CO2_cap_pro",  lowBound=0)
        self.A_EQ_cap_pro = pulp.LpVariable("Peat_avoided_EQ_cap_pro", lowBound=0)
        self.A_HH_cap_pro = pulp.LpVariable("Peat_avoided_HH_cap_pro", lowBound=0)
        self.A_EQssC_cap_pro  = pulp.LpVariable("Peat_avoided_EQssC_cap_pro",  lowBound=0)
        self.A_HHssC_cap_pro = pulp.LpVariable("Peat_avoided_HHssC_cap_pro", lowBound=0)


# linéarisation du min(., X)

##hobby
        model += self.A_CO2_cap_hobby  <= self.A_CO2_raw_hobby
        model += self.A_CO2_cap_hobby  <= self.X_CO2_hobby
        
        model += self.A_EQ_cap_hobby <= self.A_EQ_raw_hobby
        model += self.A_EQ_cap_hobby <=self.X_EQ_hobby
        
        model += self.A_HH_cap_hobby <= self.A_HH_raw_hobby
        model += self.A_HH_cap_hobby <=self.X_HH_hobby
        
        model += self.A_EQssC_cap_hobby  <= self.A_EQssC_raw_hobby
        model += self.A_EQssC_cap_hobby  <= self.X_EQssC_hobby

        model += self.A_HHssC_cap_hobby <= self.A_HHssC_raw_hobby
        model += self.A_HHssC_cap_hobby <=self.X_HHssC_hobby

##pro
        model += self.A_CO2_cap_pro  <= self.A_CO2_raw_pro
        model += self.A_CO2_cap_pro  <= self.X_CO2_pro
        
        model += self.A_EQ_cap_pro <= self.A_EQ_raw_pro
        model += self.A_EQ_cap_pro <=self.X_EQ_pro
        
        model += self.A_HH_cap_pro <= self.A_HH_raw_pro
        model += self.A_HH_cap_pro <=self.X_HH_pro
        
        model += self.A_EQssC_cap_pro  <= self.A_EQssC_raw_pro
        model += self.A_EQssC_cap_pro  <= self.X_EQssC_pro

        model += self.A_HHssC_cap_pro <= self.A_HHssC_raw_pro
        model += self.A_HHssC_cap_pro <=self.X_HHssC_pro



        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T]))+\
                                               ((self.W3/self.N3)*((trans_I3[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T3[T]))+\
                                               ((self.W4/self.N4)*((trans_I4[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T4[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))+\
                          ((self.W3/self.N3)*(self.I_inc3))+((self.W4/self.N4)*(self.I_inc4))))\
            - ((self.W/self.N)*(self.A_CO2_cap_pro+self.A_CO2_cap_hobby))- ((self.W1/self.N1) * (self.A_EQ_cap_pro+self.A_EQ_cap_hobby))\
        - ((self.W2/self.N2) * (self.A_HH_cap_pro+self.A_HH_cap_hobby)) - ((self.W3/self.N3)*(self.A_EQssC_cap_pro+self.A_EQssC_cap_hobby))\
        - ((self.W4/self.N4)*(self.A_HHssC_cap_pro+self.A_HHssC_cap_hobby))

        model.solve()
        
        #print("=== Market cap check ===")
        #print("CO2  raw:", pulp.value(self.A_CO2_raw), "cap:", pulp.value(self.A_CO2_cap), "X:", self.X_CO2)
        #print("EQssC  raw:", pulp.value(self.A_EQssC_raw), "cap:", pulp.value(self.A_EQssC_cap), "X:", self.X_EQssC)
        #print("HHssC raw:", pulp.value(self.A_HHssC_raw), "cap:", pulp.value(self.A_HHssC_cap), "X:", self.X_HHssC)
        #print("EQ raw:", pulp.value(self.A_EQ_raw), "cap:", pulp.value(self.A_EQ_cap), "X:", self.X_EQ)
        #print("HH raw:", pulp.value(self.A_HH_raw), "cap:", pulp.value(self.A_HH_cap), "X:", self.X_HH)

        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        
        print("\nFlux de coproduits vers les marchés :")
        for T in self.idtech:
            for cp in self.coprod_list:
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            if self.scale[T] == 'Centralized':
                                print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            #else:
                                #print(f"  {cp} issu de {T} → marché {mark} : {var.varValue:.2f}")
        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        total3 = 0
        total4 = 0
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans3 = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans4 = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0
                trans3 = 0
                trans4 = 0

            # traitement
            total1 += val * (trans1 + self.I_T1[T] )# - self.sub_I1[T])
            total2 += val * (trans2 + self.I_T2[T] ) #- self.sub_I2[T])
            total += val * (trans + self.I_T[T])
            total3 += val * (trans3 + self.I_T3[T])
            total4 += val * (trans4 + self.I_T4[T])# - self.sub_I[T])
        


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        total3 += self.unsort * self.I_inc3
        total4 += self.unsort * self.I_inc4
        
            # ---------- Crédit tourbe (capé par le marché) ----------
        
        # évitements bruts
#hobby market
        
        A0_raw_hobby = sum(val * (self.I_av['Peat']+ (0.7318*1)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark') 
        
        A1_raw_hobby = sum(val * (self.I_av1['Peat'] + (0.7318*0.702543)) for (T, S, cp, mark), val in self.z_opt.items()                           if mark == 'Peat_hobby_mark')

        A2_raw_hobby = sum(val *  (self.I_av2['Peat']+ (0.7318*7.1373e-6)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        A3_raw_hobby = sum(val * (self.I_av3['Peat']+ (0.7318*0.1685)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        A4_raw_hobby = sum(val * (self.I_av4['Peat']+ (0.7318*0)) for (T ,S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        
   #professionnal market     
        
        A0_raw_pro = sum(val * (self.I_av['Peat']+ (0.7318*1)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark') 
        
        A1_raw_pro = sum(val * (self.I_av1['Peat'] + (0.7318*0.702543)) for (T,S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')

        A2_raw_pro = sum(val *  (self.I_av2['Peat']+ (0.7318*7.1373e-6)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')
        
        A3_raw_pro = sum(val * (self.I_av3['Peat']+ (0.7318*0.1685)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')
        
        A4_raw_pro = sum(val * (self.I_av4['Peat']+ (0.7318*0)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')
        
        A0_cap_hobby = min(A0_raw_hobby, self.X_CO2_hobby)
        A1_cap_hobby = min(A1_raw_hobby, self.X_EQ_hobby)
        A2_cap_hobby = min(A2_raw_hobby, self.X_HH_hobby)
        A3_cap_hobby = min(A3_raw_hobby, self.X_EQssC_hobby)
        A4_cap_hobby = min(A4_raw_hobby, self.X_HHssC_hobby)
        
        A0_cap_pro = min(A0_raw_pro, self.X_CO2_pro)
        A1_cap_pro = min(A1_raw_pro, self.X_EQ_pro)
        A2_cap_pro = min(A2_raw_pro, self.X_HH_pro)
        A3_cap_pro = min(A3_raw_pro, self.X_EQssC_pro)
        A4_cap_pro = min(A4_raw_pro, self.X_HHssC_pro)
        
        total  += - A0_cap_hobby - A0_cap_pro
        total1 += - A1_cap_hobby - A1_cap_pro
        total2 += - A2_cap_hobby - A2_cap_pro
        total3  += - A3_cap_hobby - A3_cap_pro
        total4  += - A4_cap_hobby - A4_cap_pro
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}")
        print(f"Impact total {self.ind3} : {total3:.2f} {self.unit3}")
        print(f"Impact total {self.ind4} : {total4:.2f} {self.unit4}")
        
        return {self.ind : total, self.ind1: total1, self.ind2 :total2,                self.ind3 : total3 , self.ind4 : total4}
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table
    
    def get_shares_tech_site(self, decimals=2, tol=1e-9, only_feasible=True):
        """
        Retourne les % de flux par (tech, site) sur le total trié traité.
        """
        ts = {}
        total = 0.0
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            if only_feasible and self.poss_tech[T][S] == 0:
                continue
            total += float(v)
            ts[(T, S)] = ts.get((T, S), 0.0) + float(v)

        if total <= tol:
            return {}

        return {(T, S): round(100.0 * f / total, decimals) for (T, S), f in ts.items()}
    
    
        
        
        


# In[12]:




def run_batch_opt_corr_peat(scenarios_xlsx,
    municipalities,
    technologies,
    method, sort, M_hobby, M_pro,
    out_xlsx="results_batch.xlsx",
    sheet="scenarios"):
    
    scen = pd.read_excel(scenarios_xlsx, sheet_name=sheet)
    
    sort = sort

    rows = []
    for _, r in scen.iterrows():
        name = str(r["name"])
        #sort = float(r["sort"])
        #M_hobby = float(r["M_hobby"])
        #M_pro = float(r["M_pro"])

        # poids (W) : adapte si tes colonnes sont nommées autrement
        W  = float(r["W"])
        W1 = float(r["W1"])
        W2 = float(r["W2"])
        W3 = float(r["W3"])
        W4 = float(r["W4"])

        m = opt_corr_peat(
            municipalities=municipalities,
            technologies=technologies,
            method=method,
            sort=sort,
            ind="Carbon",
            ind1="EQ",
            ind2="HH",
            ind3="EQssC",
            ind4="HHssC",
            W=W, W1=W1, W2=W2, W3=W3, W4=W4,
            M_hobby = M_hobby, M_pro = M_pro
        )
        m.optimize()

        impacts = m.impact_tot()
        shares = m.get_shares_tech_site()

        rows.append({
            "name": name,
            "sort": sort,
            "impacts": impacts,
            "shares": shares,
            "idtech": list(m.idtech),
            "S_list": list(m.S_list),
            "poss_tech": m.poss_tech,
        })

    # construire tableau résultat au format techno/site (2 lignes d'en-têtes)
    df = build_output_table(rows)
    export_two_header_excel(df, out_xlsx)
    return df


# In[5]:


ARC_inv2 ='ARC_peat_corr.xlsx'
techARC_inv2 = 'tech_ARC_peat_corr.xlsx'
scen_ss_const = 'simulation_ss_constr.xlsx'
I_corrige = 'impact_brut_2 - Copie.xlsx' ##Correction sur AD-f et AD-e (surtout AD-f) pour la catégorie Carbon
I_corrige10 = 'impact_brut_2_dist_10.xlsx'
ARC_ssinv2 ='ARC_peat_corr_ss_inv.xlsx'
techARC_ssinv2 = 'tech_ARC_peat_corr_ss_inv.xlsx'
ARC_inv3 = 'ARC_peat_corr_v2.xlsx' ##decomposition du marché de la tourbe
ARC_ssinv3 = 'ARC_peat_corr_ss_inv_V2.xlsx'


# In[220]:


##without constraint - with investment


# In[233]:


ss_cons_inv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv2,
    technologies=techARC_inv2,
    method=I_corrige, sort =0.6, M_hobby = 1e20, M_pro = 1e20,
    out_xlsx="ss_const_inv.xlsx", sheet = 'Feuil1')


# In[ ]:


##without constraint - without investment


# In[227]:


ss_cons_ssinv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv2,
    technologies=techARC_ssinv2,
    method=I_corrige, sort =0.6, M_hobby = 1e20, M_pro = 1e20,
    out_xlsx="ss_const__ssinv.xlsx", sheet = 'Feuil1')


# In[ ]:


##constraint_hobby_with_investement


# In[232]:


hobby_cons_inv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv2,
    technologies=techARC_inv2,
    method=I_corrige, sort =0.6, M_hobby = 186611.81, M_pro = 1e20,
    out_xlsx="hobby_cobs_inv.xlsx", sheet = 'Feuil1')


# In[234]:


hobby_cons_ssinv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv2,
    technologies=techARC_ssinv2,
    method=I_corrige, sort =0.6, M_hobby = 186611.81, M_pro = 1e20,
    out_xlsx="hobby_const__ssinv.xlsx", sheet = 'Feuil1')


# In[235]:


pro_cons_inv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv2,
    technologies=techARC_inv2,
    method=I_corrige, sort =0.6, M_hobby = 1e20, M_pro = 185060.41,
    out_xlsx="pro_cons_inv.xlsx", sheet = 'Feuil1')


# In[236]:


pro_cons_ssinv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv2,
    technologies=techARC_ssinv2,
    method=I_corrige, sort =0.6, M_hobby = 1e20, M_pro = 185060.41,
    out_xlsx="pro_cons_ss_inv.xlsx", sheet = 'Feuil1')


# In[237]:


all_cons_inv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv2,
    technologies=techARC_inv2,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="all_cons_inv.xlsx", sheet = 'Feuil1')


# In[238]:


all_cons_ssinv = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv2,
    technologies=techARC_ssinv2,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="all_cons_ss_inv.xlsx", sheet = 'Feuil1')


# In[2]:


##toutes contraintes (excepté digestat) en décomposant le marché de tourbe (ARC_inv3 et ARCssinv3)


# In[13]:


all_cons_inv3 = run_batch_opt_corr_peat(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv3,
    technologies=techARC_inv2,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="all_cons_inv_mark_peat_decomp.xlsx", sheet = 'Feuil1')


# In[ ]:


##figures


# In[273]:


import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import re
from matplotlib.lines import Line2D
import matplotlib.patheffects as pe
import textwrap

def figure(data, sheet, title_of_graph,
             overlap_eps=1e-6,   # tolérance: 0.0 strict, 1e-6/1e-4 si arrondis
             wrap_chars=36,      # largeur max avant retour ligne dans l'encadré
             min_gap_axes=0.08   # espacement vertical entre encadrés (coords axes)
            ):

    scenario_col = "Scenario"
    indicators = ["Carbon", "EQ", "HH", "EQ (excl. CO2)", "HH (excl. CO2)"]
    inv_tag = "_inv"

    df = pd.read_excel(data, sheet)
    
    df = df[df[scenario_col].notna()]

    # conversion numérique
    for c in indicators:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # --- tri naturel (même si pas de chiffres dans les noms)
    def extract_base(name):
        return str(name).replace(inv_tag, "").strip()

    def scen_sort_key(name):
        name = str(name)
        base = extract_base(name)
        m = re.search(r"\d+", base)
        num = int(m.group()) if m else 10**9
        is_inv = inv_tag in name
        return (num, is_inv, base, name)

    df = df.sort_values(
        by=scenario_col,
        key=lambda col: col.map(scen_sort_key)
    ).reset_index(drop=True)

    # --- normalisation APRES tri
    vals = df[indicators].copy()
    vmin = vals.min(axis=0)
    vmax = vals.max(axis=0)
    denom = (vmax - vmin).replace(0, np.nan)
    vals_norm = ((vals - vmin) / denom).fillna(0.0)

    # --- liste des scénarios "base" pour couleurs (mais ici tes scénarios ne sont pas numérotés,
    # donc base == scénarios eux-mêmes si pas de _inv)
   # base_scenarios =  [
    #"minCO2",
    #"minEQ",
    #"minHH",
    #"minEQ&HH",
    #"minCO2&EQ",
    #"minCO2&HH",
    #"minCO2&EQ&HH"]
    
    bases_presentes = sorted(set(df[scenario_col].astype(str).map(extract_base)))
    distinct_13 = [...]  # ta palette
    color_map = {b: distinct_13[i % len(distinct_13)] for i, b in enumerate(bases_presentes)}


    # --- palette forte
    distinct_13 = [
        "#1f77b4",  # bleu
        "#d62728",  # rouge
        "#2ca02c",  # vert
        "#ff7f0e",  # orange
        "#9467bd",  # violet
        "#8c564b",  # brun
        "#17becf",  # cyan
        "#bcbd22",  # olive
        "#e377c2",  # rose
        "#7f7f7f",  # gris
        "#444444",  # anthracite (plus doux que noir)
        "#a65628",  # marron/orangé
        "#c2b280",  # beige foncé
    ]
    color_map = {base: distinct_13[i % len(distinct_13)] for i, base in enumerate(bases_presentes)}

    # --- interpolation
    k = 4
    x_base = np.arange(len(indicators))
    x_dense = np.linspace(x_base.min(), x_base.max(), (len(indicators) - 1) * k + 1)
    mark_positions = np.arange(0, len(x_dense), k)  # marqueurs uniquement sur les axes

    fig, ax = plt.subplots(figsize=(11, 5))

    for xi in x_base:
        ax.axvline(x=xi, color="gray", linewidth=0.8, alpha=0.35, zorder=0)

    # paramètres tracé
    lw = 1.25
    a = 1.0
    ms = 6.2
    mew = 1.6

    # ==============================
    # 1) Tracé de toutes les courbes
    # ==============================
    for i, row in df.iterrows():
        scen = str(row[scenario_col])
        base = extract_base(scen)

        y_base = vals_norm.iloc[i].values
        y_dense = np.interp(x_dense, x_base, y_base)

        is_inv = inv_tag in scen
        marker_style = "x" if is_inv else "o"
        line_style = "--" if is_inv else "-"

        line, = ax.plot(
            x_dense, y_dense,
            linestyle=line_style,
            color=color_map[base],
            linewidth=lw,
            alpha=a,
            marker=marker_style,
            markersize=ms,
            markeredgewidth=mew,
            markeredgecolor="black",
            markerfacecolor=color_map[base] if marker_style == "o" else None,
            markevery=mark_positions,
            zorder=2
        )
        
        line.set_path_effects([
            pe.Stroke(linewidth=lw + 2.2, foreground="white", alpha=0.9),
            pe.Normal()
        ])

    # ==============================================
    # 2) Détection des courbes superposées
    #    On compare les vecteurs normalisés (5 valeurs)
    # ==============================================
    Y = vals_norm.values
    n = Y.shape[0]

    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            if np.max(np.abs(Y[i] - Y[j])) <= overlap_eps:
                adj[i].append(j)
                adj[j].append(i)

    # composantes connexes = groupes superposés
    seen = [False] * n
    groups = []
    for i in range(n):
        if seen[i]:
            continue
        stack = [i]
        comp = []
        seen[i] = True
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in adj[u]:
                if not seen[v]:
                    seen[v] = True
                    stack.append(v)
        if len(comp) >= 2:
            groups.append(comp)

    # =========================================================
    # 3) Annotations avec flèche pour les groupes superposés
    #    -> on pointe vers la zone de gauche (x=0)
    #    -> encadrés empilés pour ne pas se chevaucher
    # =========================================================
    # =========================================================
# 3) Annotations alignées (flèches même longueur)
# =========================================================

    if groups:

        fig.subplots_adjust(left=0.30)

    # Position horizontale fixe des encadrés (même colonne)
        x_box_ax = -0.08   # colonne fixe (coord axes)
        x_target = x_base[0]  # on pointe vers le premier axe vertical

    # Trier du haut vers le bas
        def group_y_left(g):
            return float(np.mean([Y[idx][0] for idx in g]))

        groups_sorted = sorted(groups, key=group_y_left, reverse=True)

    # Espacement vertical régulier
        n_groups = len(groups_sorted)
        y_positions = np.linspace(0.85, 0.15, n_groups)

        for g, y_box in zip(groups_sorted, y_positions):

            scen_list = [str(df.loc[idx, scenario_col]) for idx in g]
            scen_list = sorted(scen_list, key=scen_sort_key)

            text = ", ".join(scen_list)
            text = "\n".join(textwrap.wrap(text, width=36, break_long_words=False))

            ax.annotate(
            text,
            xy=(x_target, group_y_left(g)),
            xycoords="data",
            xytext=(x_box_ax, y_box),
            textcoords=ax.transAxes,
            ha="right",   # ← alignement côté droit
            va="center",
            fontsize=8,
            color="#333333",
            bbox=dict(
                boxstyle="round,pad=0.28",
                fc="white",
                ec="#222222",
                lw=1.1,
                alpha=0.98
            ),
            arrowprops=dict(
                arrowstyle="-|>",
                lw=1.8,
                color="#333333",
                mutation_scale=18,
                shrinkA=5,
                shrinkB=5
            ),
            clip_on=False,
            zorder=10)


    # --- valeurs affichées (haut=max, bas=min) comme ton code
    for xi, ind in zip(x_base, indicators):
        ax.text(xi - 0.07, 1.02, f"{vmax[ind]:.2e}",
                ha="right", va="bottom", fontsize=8, alpha=0.85)
        ax.text(xi - 0.07, -0.02, f"{vmin[ind]:.2e}",
                ha="right", va="top", fontsize=8, alpha=0.85)

    ax.set_xticks(x_base)
    ax.set_xticklabels(indicators)
    ax.set_ylim(-0.1, 1.1)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.grid(False)
    ax.set_title(title_of_graph)

    # --- légendes (comme avant)
    color_handles = [
        Line2D([0], [0], color=color_map[base], lw=2, label=base)
        for base in bases_presentes
    ]

    style_handles = [
        Line2D([0], [0], color="black", lw=lw, marker="o",
               markersize=6, markeredgewidth=1.4, markeredgecolor="black",
               linestyle="-", label="No investment"),
        Line2D([0], [0], color="black", lw=lw, marker="x",
               markersize=6, markeredgewidth=1.6, markeredgecolor="black",
               linestyle="--", label="Authorized investment")
    ]

    leg1 = ax.legend(handles=color_handles, title="Scenarios",
                     loc="center left", bbox_to_anchor=(1.02, 0.62))
    ax.add_artist(leg1)

    ax.legend(handles=style_handles, loc="center left", bbox_to_anchor=(1.02, 0.2), title="Type")

    

    plt.tight_layout()
    plt.show()


# In[288]:


result = 'result_norm_peat_corr.xlsx'


# In[280]:


##correction digestat


# In[6]:


tech_dig_inv = 'tech_ARC_peat_dig.xlsx'
tech_dig_ss_inv = 'tech_ARC_peat_dig_ss_inv.xlsx'


# In[59]:


class opt_corr_peat2: ##contrainte de marché max  
    def __init__(self, municipalities, technologies, method, sort, ind, ind1, ind2, ind3, ind4, W, W1, W2, W3, W4,                M_hobby, M_pro):
        
        self.ind = ind
        self.sort = sort
        self.ind1= ind1
        self.ind2 = ind2
        self.ind3 = ind3
        self.ind4 = ind4
        self.W = W
        self.W1 = W1
        self.W2 = W2
        self.W3 = W3
        self.W4 = W4
        self.M_hobby = M_hobby
        self.M_pro = M_pro
        

        
        #current scenario's impact (incineration)

        
        

        
        with pd.ExcelFile(municipalities) as f:
            
            #list of municipalities
            mun = pd.read_excel(municipalities, 'mun')
            self.mun = np.array(mun['municipalities'])
            
            #total mass of waste generated
            mass = pd.read_excel(municipalities, 'waste_gen', index_col=0).fillna(0.0)
            self.tot_mass = mass['Total']['Total']
            
            self.home = mass['Home'] *self.sort
            self.build = mass['Building'] * self.sort
            self.total = mass['Total']  * self.sort
            
            self.unsort = self.tot_mass * (1 - self.sort)
            
            
            
            #possibility of a site to host a technology
            poss_tech = pd.read_excel(municipalities, 'tech_pot', index_col=0).fillna(0.0)
            self.poss_tech = poss_tech
            
            #list of potential sites
            S_list = pd.read_excel(municipalities, 'tech_pot')
            self.S_list = np.array(S_list['Site'])
            
            #collection distance of each municipality
            dist_col = pd.read_excel(municipalities, 'dist_col', index_col=0).fillna(0.0)
            self.dist_col = dist_col['dist_col']
            
            
            #distances between municipalities, potential sites and potential markets
            dist = pd.read_excel(municipalities, 'dist_market', index_col=0).fillna(0.0)
            self.dist = dist
            
            
            #list of coproduct market
            market = pd.read_excel(municipalities, 'market')
            self.market_list = np.array(market['market'])
            
            #possible market
            poss_mark = pd.read_excel(municipalities, 'market', index_col = 0)
            self.poss_mark = poss_mark
            
            #market constraint
            mark_ub = pd.read_excel(municipalities, 'mark_cons', index_col = 0)
            self.mark_ub = mark_ub['ub']

            
            
        with pd.ExcelFile(technologies) as f:
            #list of technologies
            
            type_tech = pd.read_excel(technologies, 'id_tech')
            self.idtech = np.array(type_tech['Symbol'])
            
            #scale of technology
            tech_scale = pd.read_excel(technologies, 'id_tech',index_col=0)
            self.scale = tech_scale['Type']
            
            #constraint capacity
            lb = pd.read_excel(technologies, 'tech_lb', index_col=0)
            self.lb = lb #lower bound constraint
            ub = pd.read_excel(technologies, 'tech_ub', index_col=0)
            self.ub = ub  #upper bound constraint
            
            
            #list of possible coproduct
            coprod_list = pd.read_excel(technologies, 'subs_ratio')
            self.coprod_list = np.array(coprod_list['Service'])
            
            #yield of coproduct production for each technology
            coprod = pd.read_excel(technologies, 'coproduct', index_col=0)
            self.coprod = coprod
            

            
            
            #list of possible conventional product 
            #c_p_list = pd.read_excel(technologies, 'subs_ratio')
            
            c_p_list = pd.read_excel(technologies, 'avoid')
            self.c_p_list = np.array(c_p_list['Conventional product'])
            
            #substitution ratio conventional product/co-product
            #sub_ratio = pd.read_excel(technologies, 'subs_ratio', index_col=0)
            #self.sub_ratio = sub_ratio
            
            #avoided product for each technology
            avoid = pd.read_excel(technologies, 'avoid',index_col=0)
            self.avoid = avoid

                
        
        with pd.ExcelFile(method) as f:
            imp = pd.read_excel(method, 'Feuil1')
            
            #unit
            self.unit = imp[self.ind][0]
            self.unit1 = imp[self.ind1][0]
            self.unit2 = imp[self.ind2][0]
            self.unit3 = imp[self.ind3][0]
            self.unit4 = imp[self.ind4][0]
            
            
            #impacts of processes
            ##impact of transport (after collection, from the municipality to the treatment site)
            self.tr_w = imp[self.ind][1]
            self.tr_w1 = imp[self.ind1][1]
            self.tr_w2 = imp[self.ind2][1]
            self.tr_w3 = imp[self.ind3][1]
            self.tr_w4 = imp[self.ind4][1]
            
            ##impact of collection
            self.col_w = imp[self.ind][2]
            self.col_w1 = imp[self.ind1][2]
            self.col_w2 = imp[self.ind2][2]
            self.col_w3 = imp[self.ind3][2]
            self.col_w4 = imp[self.ind4][2]
            
            ##impact of other processes
            imp["Step"] = imp['Process'].where(imp["Unit reference"].isna())  # les titres ont Unit reference vide
            imp["Step"] = imp["Step"].ffill()
            
            ###treatment technology
            tech = imp[imp["Step"] == "Treatment technologies"]
            tech = tech.set_index('Process')
            self.I_T = tech[self.ind]
            self.I_T1 = tech[self.ind1]
            self.I_T2 = tech[self.ind2]
            self.I_T3 = tech[self.ind3]
            self.I_T4 = tech[self.ind4]
            
            self.I_inc = self.I_T['Inc']
            self.I_inc1 = self.I_T1['Inc']
            self.I_inc2 = self.I_T2['Inc']
            self.I_inc3 = self.I_T3['Inc']
            self.I_inc4 = self.I_T4['Inc']
            
            ###Impact coproduct transport
            tr_cop = imp[imp["Step"] == "Transport of coproduct"]
            tr_cop = tr_cop.set_index('Process')
            
            self.I_tr_cop = tr_cop[self.ind]
            self.I_tr_cop1 = tr_cop[self.ind1]
            self.I_tr_cop2 = tr_cop[self.ind2]
            self.I_tr_cop3 = tr_cop[self.ind3]
            self.I_tr_cop4 = tr_cop[self.ind4]
            
            ###Avoided impact related to conventional product production
            av = imp[imp["Step"] == "Conventional products"]
            av = av.set_index('Process')
            
            self.I_av = av[self.ind]
            self.I_av1 = av[self.ind1]
            self.I_av2 = av[self.ind2]
            self.I_av3 = av[self.ind3]
            self.I_av4 = av[self.ind4]
            
            ###Net impact of coproduct use
            u_cop = imp[imp["Step"] == 'Net impact of coproduct use']
            u_cop = u_cop.set_index('Process')
            
            self.I_u_cop = u_cop[self.ind]
            self.I_u_cop1 = u_cop[self.ind1]
            self.I_u_cop2 = u_cop[self.ind2]
            self.I_u_cop3 = u_cop[self.ind3]
            self.I_u_cop4 = u_cop[self.ind4]
            
        self.N = self.tot_mass * self.I_inc
        self.N1 = self.tot_mass * self.I_inc1
        self.N2 = self.tot_mass * self.I_inc2
        self.N3 = self.tot_mass * self.I_inc3
        self.N4 = self.tot_mass * self.I_inc4
        
    
    def optimize(self):
        model = pulp.LpProblem("biowaste_management", pulp.LpMinimize)
        
        self.origin = ['Home','Building']
        
        #variables
        x = {} #quantity of waste sent to each technology
        y = {} #activation or not of a technology
        z = {} #quantity of coproduct sent to market segment
        
        for M in self.mun:
            for O in self.origin:
                for T in self.idtech:
                    for S in self.S_list:
                        x[(M, O, T, S)] = pulp.LpVariable(f"x_{M}_{O}_{T}_{S}", lowBound=0)
        
        for T in self.idtech:
            for S in self.S_list:
                y[(T,S)] = pulp.LpVariable(f"y_{T}_{S}", cat="Binary")
        
        for T in self.idtech:
            for S in self.S_list:
                for cp in self.coprod_list:
                    for mark in self.market_list:
                        z[(T,S,cp,mark)] = pulp.LpVariable(f"z_{T}_{S}_{cp}_{mark}", lowBound = 0)
        

        
        #constraints
        ##all waste has to be treated
        for M in self.mun:
            model += pulp.lpSum(x[(M, 'Home', T, S)] for T in self.idtech for S in self.S_list) == self.home[M]
            model += pulp.lpSum(x[(M, 'Building', T, S)] for T in self.idtech for S in self.S_list) == self.build[M]
        
        ##HC can not be used in buildings
        for M in self.mun:
            model += pulp.lpSum(x[M, 'Building', 'HC',S] for S in self.S_list) == 0
        
        ##constraint capacity
        for T in self.idtech:
            for S in self.S_list:
                if not pd.isna(self.lb[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) >= self.lb[T][S]*y[(T,S)]
                if not pd.isna(self.ub[T][S]):
                    model += pulp.lpSum(x[(M, O, T, S)] for M in self.mun for O in self.origin) <= self.ub[T][S]*y[(T,S)]
        
        #self.transport = 2e-1 #kgCO2/t.km
        
        ##a technology can be installed only at a site which can host it
        for T in self.idtech:
            for S in self.S_list:
                if self.poss_tech[T][S] == 0:
                    model += y[(T,S)] == 0
                    model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin for M in self.mun) == 0
        
        
        ##decentralized technology can receive only waste from the site where it is installed
        #(this constraint may be unnecessary)
        for M in self.mun:
            for S in self.S_list:
                for T in self.idtech:
                    if self.scale[T] == 'Decentralized':
                        if M != S :
                            model += (y[(T,S)]) == 0
                            model +=  pulp.lpSum(x[(M, O, T, S)] for O in self.origin) == 0  
                            #--> this last line seems to mean that decentralized process should not exist ...

        

        
        
        #impact of transport of waste to the facility (collection+transpot to facility)
        trans_I1 = {}
        trans_I2 = {}
        trans_I = {}
        trans_I3 = {}
        trans_I4 = {}
        
        
        #for T in self.idtech:
        for M in self.mun:
            for S in self.S_list:
                trans_I1[(M,S)] = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans_I2[(M,S)] = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans_I[(M,S)] = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans_I3[(M,S)] = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans_I4[(M,S)] = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
        
  
        

                        
        #impact of transport of coproduct


        
        
 

        ##all coproduct have to find its market.
        for cp in self.coprod_list:
            for T in self.idtech:
                for S in self.S_list:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for mark in self.market_list) ==                     pulp.lpSum(x[(M,O,T,S)]* self.coprod[T][cp] for M in self.mun                            for O in self.origin)
                    
        ##constraint : coproduct should be sent to a market which can valorize it.
        for cp in self.coprod_list:
            for mark in self.market_list:
                if self.poss_mark[cp][mark] == 0:
                    model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list) ==0
                    
        ##constraint market of coproduct
        for mark in self.market_list:
            if not pd.isna(self.mark_ub[mark]):
                model += pulp.lpSum(z[(T,S,cp,mark)] for T in self.idtech for S in self.S_list                                   for cp in self.coprod_list) <= self.mark_ub[mark]
                    

        

# évitements bruts
#hobby market
        
        self.A_CO2_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av['Peat']+ (0.7318*1))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_EQ_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av1['Peat']+ (0.7318*0.702543))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)

        self.A_HH_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av2['Peat']+ (0.7318*7.1373e-6))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_EQssC_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av3['Peat']+ (0.7318*0.1685))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        self.A_HHssC_raw_hobby = pulp.lpSum(z[(T, S, cp, 'Peat_hobby_mark')] * (self.I_av4['Peat']+ (0.7318*0))
    for cp in self.coprod_list for T in self.idtech for S in self.S_list)
        
        
   #professionnal market     
        
        self.A_CO2_raw_pro = pulp.lpSum(z[(T, S,cp, 'Peat_pro_mark')] * (self.I_av['Peat']+ (0.7318*1))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_EQ_raw_pro = pulp.lpSum(z[(T, S,cp, 'Peat_pro_mark')] * (self.I_av1['Peat']+ (0.7318*0.702543))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)

        self.A_HH_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av2['Peat']+ (0.7318*7.1373e-6))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_EQssC_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av3['Peat']+ (0.7318*0.1685))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)
        
        self.A_HHssC_raw_pro = pulp.lpSum(z[(T, S, cp, 'Peat_pro_mark')] * (self.I_av4['Peat']+ (0.7318*0))
    for T in self.idtech for S in self.S_list for cp in self.coprod_list)




# linéarisation du min(., X)





        #objective
        
        model +=  pulp.lpSum(x[(M, O, T, S)] * (((self.W1/self.N1)*((trans_I1[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T1[T]))+                                                ((self.W2/self.N2)*((trans_I2[(M,S)] if self.scale[T] == 'Centralized' else 0.0)                                                                    + self.I_T2[T]))
                                               +((self.W/self.N)*((trans_I[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T[T]))+\
                                               ((self.W3/self.N3)*((trans_I3[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T3[T]))+\
                                               ((self.W4/self.N4)*((trans_I4[(M,S)] if self.scale[T] == 'Centralized' else 0.0)\
                                                                    + self.I_T4[T])))\
                             for M in self.mun for O in self.origin for T in self.idtech for S in self.S_list) \
        +  (self.unsort * (((self.W1/self.N1)*(self.I_inc1))+((self.W2/self.N2)*(self.I_inc2))+((self.W/self.N)*(self.I_inc))+\
                          ((self.W3/self.N3)*(self.I_inc3))+((self.W4/self.N4)*(self.I_inc4))))\
            - ((self.W/self.N)*(self.A_CO2_raw_pro+self.A_CO2_raw_hobby))- ((self.W1/self.N1) * (self.A_EQ_raw_pro+self.A_EQ_raw_hobby))\
        - ((self.W2/self.N2) * (self.A_HH_raw_pro+self.A_HH_raw_hobby)) - ((self.W3/self.N3)*(self.A_EQssC_raw_pro+self.A_EQssC_raw_hobby))\
        - ((self.W4/self.N4)*(self.A_HHssC_raw_pro+self.A_HHssC_raw_hobby))

        model.solve()
        
        #print("=== Market cap check ===")
        #print("CO2  raw:", pulp.value(self.A_CO2_raw), "cap:", pulp.value(self.A_CO2_cap), "X:", self.X_CO2)
        #print("EQssC  raw:", pulp.value(self.A_EQssC_raw), "cap:", pulp.value(self.A_EQssC_cap), "X:", self.X_EQssC)
        #print("HHssC raw:", pulp.value(self.A_HHssC_raw), "cap:", pulp.value(self.A_HHssC_cap), "X:", self.X_HHssC)
        #print("EQ raw:", pulp.value(self.A_EQ_raw), "cap:", pulp.value(self.A_EQ_cap), "X:", self.X_EQ)
        #print("HH raw:", pulp.value(self.A_HH_raw), "cap:", pulp.value(self.A_HH_cap), "X:", self.X_HH)

        
        #stockage des résultats
        self.x_opt = {k: v.varValue for k, v in x.items() if v.varValue is not None}
        self.y_opt = {k: v.varValue for k, v in y.items() if v.varValue is not None}
        self.z_opt = {k: v.varValue for k, v in z.items() if v.varValue is not None}

        
        #print("\n=== Valeurs des variables d’activation y(T,S) ===")
        
        for T in self.idtech:
            for S in self.S_list:
                val = y[(T, S)].varValue
                #print(f"y[{T},{S}] = {val}")

        
        #for M in self.mun:
            #print(f"\nCommune : {M}")
            #for T in self.idtech:
            #    for O in self.origin:
             #       for S in self.S_list:
              #          var = x[(M, O, T,S)]
               #         if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            #print(f"  {O} → {T} in {S} : {var.varValue:.1f} tonnes")
                            #else:
                                #print(f"  {O} → {T} : {var.varValue:.1f} tonnes")
        print("\nRésultats:")
        
        print("\nFlux de coproduits vers les marchés :")
        
        totals = {}
        for T in self.idtech:
            for cp in self.coprod_list:
                if cp not in ["Peat_dec", "Peat_ind"]:
                    continue             
                for mark in self.market_list:
                    for S in self.S_list:
                        var = z[(T,S, cp, mark)]
                        if var.varValue is not None and var.varValue > 1e-3:
                            #if self.scale[T] == 'Centralized':
                            
                            flux = float(var.varValue)
                            print(f"  {cp} issu de {T} installé dans {S} → marché {mark} : {var.varValue:.2f}")
                            
                            key = (cp, mark)
                            totals[key]= totals.get(key, 0) + flux
        print ("\nSomme totale vers chaque marché:")
        
        for (cp,mark), total in totals.items():
            print(f"{cp} vers {mark} : {total:.2f}")
                            

        print("\nTechnologies activées et flux associés :")

        for T in self.idtech:
            for S in self.S_list:
                if y[(T, S)].varValue == 1:
                    total_TS = sum(x[(M, O, T, S)].varValue for M in self.mun for O in self.origin                                   if x[(M, O, T, S)].varValue is not None)
                    print(f"  {T} installé en {S} : {total_TS:.1f} tonnes")

        
                        
        
        print(f"\nStatut de l'optimisation : {pulp.LpStatus[model.status]}")
        
        self.tot_imp = pulp.value(model.objective) 
                                                                      
        print(f"Impact environnemental total minimal : {(self.tot_imp):.2f}")
        
    
    def impact_tot(self):
        
        total = 0
        total1 = 0
        total2 = 0
        total3 = 0
        total4 = 0
        
    #Calcule les impacts environnementaux pour plusieurs catégories en utilisant la solution optimale déjà trouvée.

        # ---------- Impact traitement + transport des déchets ----------
        for (M, O, T, S), val in self.x_opt.items():
# transport déchets
            if self.scale[T] == 'Centralized':
                trans1 = (self.col_w1 * self.dist_col[M]) + (self.tr_w1 * self.dist[M][S])
                trans2 = (self.col_w2 * self.dist_col[M]) + (self.tr_w2 * self.dist[M][S])
                trans = (self.col_w * self.dist_col[M]) + (self.tr_w * self.dist[M][S])
                trans3 = (self.col_w3 * self.dist_col[M]) + (self.tr_w3 * self.dist[M][S])
                trans4 = (self.col_w4 * self.dist_col[M]) + (self.tr_w4 * self.dist[M][S])
                
            else:
                trans1 = 0.0
                trans2 = 0
                trans = 0
                trans3 = 0
                trans4 = 0
            
            #transport

            total1 += val * (trans1 )# - self.sub_I1[T])
            total2 += val * (trans2) #- self.sub_I2[T])
            total += val * (trans )
            total3 += val * (trans3 )
            total4 += val * (trans4 )# - self.sub_I[T])
            
        print(total1)
        print(total2)
        print(total)
        
        for (M, O, T, S), val in self.x_opt.items():
            
            # traitement 
            total1 += val * (self.I_T1[T] )# - self.sub_I1[T])
            total2 += val * (self.I_T2[T] ) #- self.sub_I2[T])
            total += val * (self.I_T[T])
            total3 += val * (self.I_T3[T])
            total4 += val * ( self.I_T4[T])# - self.sub_I[T])
            
        


        # ---------- Incinération des déchets non triés ----------
        total1 += self.unsort * self.I_inc1
        total2 += self.unsort * self.I_inc2
        total += self.unsort * self.I_inc
        total3 += self.unsort * self.I_inc3
        total4 += self.unsort * self.I_inc4
        
            # ---------- Crédit tourbe (capé par le marché) ----------
        
        # évitements bruts
#hobby market
        
        A0_raw_hobby = sum(val * (self.I_av['Peat']+ (0.7318*1)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark') 
        
        A1_raw_hobby = sum(val * (self.I_av1['Peat'] + (0.7318*0.702543)) for (T, S, cp, mark), val in self.z_opt.items()                           if mark == 'Peat_hobby_mark')

        A2_raw_hobby = sum(val *  (self.I_av2['Peat']+ (0.7318*7.1373e-6)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        A3_raw_hobby = sum(val * (self.I_av3['Peat']+ (0.7318*0.1685)) for (T, S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        A4_raw_hobby = sum(val * (self.I_av4['Peat']+ (0.7318*0)) for (T ,S, cp, mark), val in self.z_opt.items()                          if mark == 'Peat_hobby_mark')
        
        
   #professionnal market     
        
        A0_raw_pro = sum(val * (self.I_av['Peat']+ (0.7318*1)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark') 
        
        A1_raw_pro = sum(val * (self.I_av1['Peat'] + (0.7318*0.702543)) for (T,S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')

        A2_raw_pro = sum(val *  (self.I_av2['Peat']+ (0.7318*7.1373e-6)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')
        
        A3_raw_pro = sum(val * (self.I_av3['Peat']+ (0.7318*0.1685)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')
        
        A4_raw_pro = sum(val * (self.I_av4['Peat']+ (0.7318*0)) for (T, S, cp, mark), val in self.z_opt.items()                        if mark == 'Peat_pro_mark')

        total  += - A0_raw_hobby - A0_raw_pro
        total1 += - A1_raw_hobby - A1_raw_pro
        total2 += - A2_raw_hobby - A2_raw_pro
        total3  += - A3_raw_hobby - A3_raw_pro
        total4  += - A4_raw_hobby - A4_raw_pro
        

        print(f"Impact total {self.ind1} : {total1:.2f} {self.unit1} , Transport: {trans1:.2e} {self.unit1}")
        print(f"Impact total {self.ind2} : {total2:.2f} {self.unit2}, Transport: {trans2:.2e} {self.unit2}")
        print(f"Impact total {self.ind} : {total:.2f} {self.unit}, Transport: {trans:.2e} {self.unit}")
        print(f"Impact total {self.ind3} : {total3:.2f} {self.unit3}, Transport: {trans3:.2e} {self.unit3}")
        print(f"Impact total {self.ind4} : {total4:.2f} {self.unit4}, Transport: {trans4:.2e} {self.unit4}")
        
        return {self.ind : total, self.ind1: total1, self.ind2 :total2,                self.ind3 : total3 , self.ind4 : total4}
    
    
    def flows_all(self, tol=1e-6):
        rows = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows.append({"Commune": M, "Origin": O, "Tech": T, "Site": S, "Flow_t": float(v)})
        return pd.DataFrame(rows)

    def flows_techno(self, tol=1e-6):
        rows_T = []
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            rows_T.append({"Tech": T, "Site": S, "Flux_t": float(v)})

        df = pd.DataFrame(rows_T)

    # agrégation
        table = (df.groupby(["Tech", "Site"], as_index=False)["Flux_t"]
        .sum()
        .sort_values("Flux_t", ascending=False)
        .reset_index(drop=True)
    )

        total = table["Flux_t"].sum()
        table["Pourcentage_%"] = 100 * table["Flux_t"] / total

        
        print(table)

        return table
    
    def get_shares_tech_site(self, decimals=2, tol=1e-9, only_feasible=True):
        """
        Retourne les % de flux par (tech, site) sur le total trié traité.
        """
        ts = {}
        total = 0.0
        for (M, O, T, S), v in self.x_opt.items():
            if v is None or v <= tol:
                continue
            if only_feasible and self.poss_tech[T][S] == 0:
                continue
            total += float(v)
            ts[(T, S)] = ts.get((T, S), 0.0) + float(v)

        if total <= tol:
            return {}

        return {(T, S): round(100.0 * f / total, decimals) for (T, S), f in ts.items()}
    
    
        
        
        


# In[60]:




def run_batch_opt_corr_peat2(scenarios_xlsx,
    municipalities,
    technologies,
    method, sort, M_hobby, M_pro,
    out_xlsx="results_batch.xlsx",
    sheet="scenarios"):
    
    scen = pd.read_excel(scenarios_xlsx, sheet_name=sheet)
    
    sort = sort

    rows = []
    for _, r in scen.iterrows():
        name = str(r["name"])
        #sort = float(r["sort"])
        #M_hobby = float(r["M_hobby"])
        #M_pro = float(r["M_pro"])

        # poids (W) 
        W  = float(r["W"])
        W1 = float(r["W1"])
        W2 = float(r["W2"])
        W3 = float(r["W3"])
        W4 = float(r["W4"])

        m = opt_corr_peat2(
            municipalities=municipalities,
            technologies=technologies,
            method=method,
            sort=sort,
            ind="Carbon",
            ind1="EQ",
            ind2="HH",
            ind3="EQssC",
            ind4="HHssC",
            W=W, W1=W1, W2=W2, W3=W3, W4=W4,
            M_hobby = M_hobby, M_pro = M_pro
        )
        m.optimize()

        impacts = m.impact_tot()
        shares = m.get_shares_tech_site()

        rows.append({
            "name": name,
            "sort": sort,
            "impacts": impacts,
            "shares": shares,
            "idtech": list(m.idtech),
            "S_list": list(m.S_list),
            "poss_tech": m.poss_tech,
        })
        
        
        #origin_flow = m.flows_all()
        
        #print (origin_flow)

    # construire tableau résultat au format techno/site (2 lignes d'en-têtes)
    df = build_output_table(rows)
    export_two_header_excel(df, out_xlsx)
    return df


# In[61]:


all_dig_inv2 = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="all_dig_inv2_VF.xlsx", sheet = 'Feuil1')


# In[62]:


all_dig_ssinv2 = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="all_dig_ssinv2.xlsx", sheet = 'Feuil1')


# In[38]:


pareto = 'simulation_pareto.xlsx'


# In[39]:


#simulation avec contraintes


# In[40]:


#SIMULATION AVEC INVESTISSEMENT


# In[41]:


#CO2-EQ pareto


# In[42]:


par1_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="CO2_EQ_VF.xlsx", sheet = 'par_CO2_EQ')


# In[189]:


#CO2-HH pareto


# In[43]:


par2_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="CO2_HH_VF.xlsx", sheet = 'par_CO2_HH')


# In[178]:


#EQ-HH


# In[199]:


par3_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="EQ_HH.xlsx", sheet = 'par_EQ_HH')


# In[180]:


#SIMULATION SANS INVESTISSEMENT


# In[191]:


#CO2-EQ pareto


# In[192]:


par1_ss_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="CO2_EQ_ss_inv.xlsx", sheet = 'par_CO2_EQ')


# In[193]:


par2_ss_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="CO2_HH_ss_inv.xlsx", sheet = 'par_CO2_HH')


# In[201]:


par3_ss_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="EQ_HH_ss_inv.xlsx", sheet = 'par_EQ_HH')


# In[ ]:


#simulation sans contrainte


# In[ ]:


#avec investissement


# In[195]:


par1_sscont_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv_ss_cons,
    technologies=tech_ss_cons,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="CO2_EQ_ss_cons_inv.xlsx", sheet = 'par_CO2_EQ')


# In[196]:


par2_sscont_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv_ss_cons,
    technologies=tech_ss_cons,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="CO2_HH_ss_cons_inv.xlsx", sheet = 'par_CO2_HH')


# In[ ]:


#sans investissement


# In[197]:


par1_sscont_ss_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ss_inv_ss_cons,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="CO2_EQ_ss_cons_ssinv.xlsx", sheet = 'par_CO2_EQ')


# In[198]:


par2_sscont_ss_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ss_inv_ss_cons,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="CO2_HH_ss_cons_ssinv.xlsx", sheet = 'par_CO2_HH')


# In[84]:


##analyse de sensibilité par rapport au taux de tri


# In[ ]:


#avec investissement


# In[44]:


AS_pos_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.9, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="ASinv_tri_pos_VF.xlsx", sheet = 'AS')


# In[45]:


AS_neg_inv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.3, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="ASinv_tri_neg_VF.xlsx", sheet = 'AS')


# In[206]:


AS_pos_ssinv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.9, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="ASssinv_atri_pos.xlsx", sheet = 'AS')


# In[207]:


AS_neg_ssinv = run_batch_opt_corr_peat2(scenarios_xlsx=pareto,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.3, M_hobby = 186611.81 , M_pro = 185060.41,
    out_xlsx="ASssinv_tri_neg.xlsx", sheet = 'AS')


# In[ ]:


##AS par rapport aux contraintes de marché


# In[85]:


AS_ssinv_neg_hobby = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = (186611.81*(1-0.1)) , M_pro = 185060.41,
    out_xlsx="AS_neg_ssinv_hobby.xlsx", sheet = 'Feuil1')


# In[89]:


AS_ssinv_pos_hobby = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ssinv3,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = (186611.81*(1+0.1)) , M_pro = 185060.41,
    out_xlsx="AS_pos_ssinv_hobby.xlsx", sheet = 'Feuil1')


# In[90]:


AS_pos_inv_hobby = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = (186611.81*(1+0.1)) , M_pro = 185060.41,
    out_xlsx="AS_pos_inv_hobby.xlsx", sheet = 'Feuil1')


# In[91]:


AS_neg_inv_hobby = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv3,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = (186611.81*(1-0.1)) , M_pro = 185060.41,
    out_xlsx="AS_neg_inv_hobby.xlsx", sheet = 'Feuil1')


# In[92]:


##Refaire le scénario sans contrainte en considérant déjà que le digestat ne substitue pas de la tourbe


# In[12]:


ARC_inv_ss_cons = 'ARC_inv_ss_cons_VF.xlsx'
ARC_ss_inv_ss_cons = 'ARC_ss_inv_ss_cons_VF.xlsx'


# In[100]:


all_dig_inv_vf = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv_ss_cons,
    technologies=tech_dig_inv,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="all_dig_inv_vf.xlsx", sheet = 'Feuil1')


# In[101]:


all_dig_ssinv_vf = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_ss_inv_ss_cons,
    technologies=tech_dig_ss_inv,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="all_dig_ssinv_vf.xlsx", sheet = 'Feuil1')


# In[ ]:


##si on supprime les contraintes technologiques pour les nouveux investissements


# In[13]:


tech_ss_cons = 'tech_ARC_peat_dig_ss_cont_tech.xlsx'


# In[14]:


#sans contrainte de marché, sans contrainte de capacité tech


# In[28]:


simul_ss_tec_ss_mark_inv = run_batch_opt_corr_peat2(scenarios_xlsx=scen_ss_const,
    municipalities= ARC_inv_ss_cons,
    technologies=tech_ss_cons,
    method=I_corrige, sort =0.6, M_hobby = 1e20 , M_pro = 1e20,
    out_xlsx="ss_cons_mark_tech_inv_VF.xlsx", sheet = 'Feuil1')


# In[105]:


##Figures de l'article


# In[31]:


import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import re
from matplotlib.lines import Line2D
import matplotlib.patheffects as pe
import textwrap

def figure_paper(data, sheet, title_of_graph,
             overlap_eps=1e-6,   # tolérance: 0.0 strict, 1e-6/1e-4 si arrondis
             wrap_chars=36,      # largeur max avant retour ligne dans l'encadré
             min_gap_axes=0.08   # espacement vertical entre encadrés (coords axes)
            ):

    scenario_col = "Scenario"
    indicators = ["CC (kgCO2 eq)", "EQ (PDF.m2.yr)", "HH (DALY)"]
    inv_tag = "_inv"

    df = pd.read_excel(data, sheet)

    # conversion numérique
    for c in indicators:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # --- tri naturel (même si pas de chiffres dans les noms)
    def extract_base(name):
        return str(name).replace(inv_tag, "")

    def scen_sort_key(name):
        name = str(name)
        base = extract_base(name)
        m = re.search(r"\d+", base)
        num = int(m.group()) if m else 10**9
        is_inv = inv_tag in name
        return (num, is_inv, base, name)

    df = df.sort_values(
        by=scenario_col,
        key=lambda col: col.map(scen_sort_key)
    ).reset_index(drop=True)

    # --- normalisation APRES tri
    vals = df[indicators].copy()
    vmin = vals.min(axis=0)
    vmax = vals.max(axis=0)
    denom = (vmax - vmin).replace(0, np.nan)
    vals_norm = ((vals - vmin) / denom).fillna(0.0)

    # --- liste des scénarios "base" pour couleurs (mais ici tes scénarios ne sont pas numérotés,
    # donc base == scénarios eux-mêmes si pas de _inv)
    base_scenarios =  [
    "minCC",
    "minEQ",
    "minHH"]


    # --- palette forte
    distinct_13 = [
        "#1f77b4",  # bleu
        "#d62728",  # rouge
        "#2ca02c",  # vert
        "#ff7f0e",  # orange
        "#9467bd",  # violet
        "#8c564b",  # brun
        "#17becf",  # cyan
        "#bcbd22",  # olive
        "#e377c2",  # rose
        "#7f7f7f",  # gris
        "#444444",  # anthracite (plus doux que noir)
        "#a65628",  # marron/orangé
        "#c2b280",  # beige foncé
    ]
    color_map = {base: distinct_13[i % len(distinct_13)] for i, base in enumerate(base_scenarios)}

    # --- interpolation
    k = 4
    x_base = np.arange(len(indicators))
    x_dense = np.linspace(x_base.min(), x_base.max(), (len(indicators) - 1) * k + 1)
    mark_positions = np.arange(0, len(x_dense), k)  # marqueurs uniquement sur les axes

    fig, ax = plt.subplots(figsize=(11, 5))

    for xi in x_base:
        ax.axvline(x=xi, color="gray", linewidth=0.8, alpha=0.35, zorder=0)

    # paramètres tracé
    lw = 1.25
    a = 1.0
    ms = 6.2
    mew = 1.6

    # ==============================
    # 1) Tracé de toutes les courbes
    # ==============================
    for i, row in df.iterrows():
        scen = str(row[scenario_col])
        base = extract_base(scen)

        y_base = vals_norm.iloc[i].values
        y_dense = np.interp(x_dense, x_base, y_base)

        is_inv = inv_tag in scen
        marker_style = "x" if is_inv else "o"
        line_style = "--" if is_inv else "-"

        line, = ax.plot(
            x_dense, y_dense,
            linestyle=line_style,
            color=color_map[base],
            linewidth=lw,
            alpha=a,
            marker=marker_style,
            markersize=ms,
            markeredgewidth=mew,
            markeredgecolor="black",
            markerfacecolor=color_map[base] if marker_style == "o" else None,
            markevery=mark_positions,
            zorder=2
        )

        line.set_path_effects([
            pe.Stroke(linewidth=lw + 2.2, foreground="white", alpha=0.9),
            pe.Normal()
        ])

    # ==============================================
    # 2) Détection des courbes superposées
    #    On compare les vecteurs normalisés (5 valeurs)
    # ==============================================
    Y = vals_norm.values
    n = Y.shape[0]

    adj = [[] for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            if np.max(np.abs(Y[i] - Y[j])) <= overlap_eps:
                adj[i].append(j)
                adj[j].append(i)

    # composantes connexes = groupes superposés
    seen = [False] * n
    groups = []
    for i in range(n):
        if seen[i]:
            continue
        stack = [i]
        comp = []
        seen[i] = True
        while stack:
            u = stack.pop()
            comp.append(u)
            for v in adj[u]:
                if not seen[v]:
                    seen[v] = True
                    stack.append(v)
        if len(comp) >= 2:
            groups.append(comp)

    # =========================================================
    # 3) Annotations avec flèche pour les groupes superposés
    #    -> on pointe vers la zone de gauche (x=0)
    #    -> encadrés empilés pour ne pas se chevaucher
    # =========================================================
    # =========================================================
# 3) Annotations alignées (flèches même longueur)
# =========================================================

    if groups:

        fig.subplots_adjust(left=0.30)

    # Position horizontale fixe des encadrés (même colonne)
        x_box_ax = -0.08   # colonne fixe (coord axes)
        x_target = x_base[0]  # on pointe vers le premier axe vertical

    # Trier du haut vers le bas
        def group_y_left(g):
            return float(np.mean([Y[idx][0] for idx in g]))

        groups_sorted = sorted(groups, key=group_y_left, reverse=True)

    # Espacement vertical régulier
        n_groups = len(groups_sorted)
        y_positions = np.linspace(0.85, 0.15, n_groups)

        for g, y_box in zip(groups_sorted, y_positions):

            scen_list = [str(df.loc[idx, scenario_col]) for idx in g]
            scen_list = sorted(scen_list, key=scen_sort_key)

            text = ", ".join(scen_list)
            text = "\n".join(textwrap.wrap(text, width=36, break_long_words=False))

            ax.annotate(
            text,
            xy=(x_target, group_y_left(g)),
            xycoords="data",
            xytext=(x_box_ax, y_box),
            textcoords=ax.transAxes,
            ha="right",   # ← alignement côté droit
            va="center",
            fontsize=8,
            color="#333333",
            bbox=dict(
                boxstyle="round,pad=0.28",
                fc="white",
                ec="#222222",
                lw=1.1,
                alpha=0.98
            ),
            arrowprops=dict(
                arrowstyle="-|>",
                lw=1.8,
                color="#333333",
                mutation_scale=18,
                shrinkA=5,
                shrinkB=5
            ),
            clip_on=False,
            zorder=10)


    # --- valeurs affichées (haut=max, bas=min) comme ton code
    for xi, ind in zip(x_base, indicators):
        ax.text(xi - 0.07, 1.02, f"{vmax[ind]:.2e}",
                ha="right", va="bottom", fontsize=12, alpha=1) #fontsize=8, alpha=0.85
        ax.text(xi - 0.07, -0.02, f"{vmin[ind]:.2e}",
                ha="right", va="top", fontsize=12, alpha=1) #fontsize=8, alpha=0.85

    ax.set_xticks(x_base)
    ax.set_xticklabels(indicators)
    ax.set_ylim(-0.1, 1.1)
    ax.set_yticks([])
    ax.spines["left"].set_visible(False)
    ax.grid(False)
    ax.set_title(title_of_graph)

    # --- légendes (comme avant)
    color_handles = [
        Line2D([0], [0], color=color_map[base], lw=2, label=base)
        for base in base_scenarios
    ]

    style_handles = [
        Line2D([0], [0], color="black", lw=lw, marker="o",
               markersize=6, markeredgewidth=1.4, markeredgecolor="black",
               linestyle="-", label="No investment"),
        Line2D([0], [0], color="black", lw=lw, marker="x",
               markersize=6, markeredgewidth=1.6, markeredgecolor="black",
               linestyle="--", label="Authorized investment")
    ]

    leg1 = ax.legend(handles=color_handles, title="Scenarios",
                     loc="center left", bbox_to_anchor=(1.02, 0.62))
    ax.add_artist(leg1)

    ax.legend(handles=style_handles, loc="center left", bbox_to_anchor=(1.02, 0.2), title="Type")

    plt.tight_layout()
    plt.show()


# In[32]:


#Figure 1 : Mono-objectif : sans contrainte et avec contrainte (sans inv et avec inv)


# In[33]:


#sans contrainte


# In[34]:


fig1 = 'figure1_mono_obj_VF.xlsx'


# In[35]:


fig1_ss_cons = figure_paper(fig1, 'without_cons', '')


# In[36]:


fig1_cons = figure_paper(fig1, 'with_cons', '')


# In[ ]:




